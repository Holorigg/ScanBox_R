import sys
import os
from pathlib import Path
import re
import json
import csv
from datetime import datetime
from time import time
import random
import threading
from collections import deque

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QLabel, QLineEdit,
    QPushButton, QVBoxLayout, QHBoxLayout, QGridLayout, QGroupBox,
    QMessageBox, QFileDialog, QInputDialog, QTextEdit,
    QTreeWidget, QTreeWidgetItem, QMenu, QAction, QHeaderView,
    QCheckBox, QScrollArea, QMenuBar,
    QDialog, QSpacerItem, QSizePolicy, QComboBox,
    QDialogButtonBox, QFrame, QTableWidget, QTableWidgetItem, QAbstractItemView,
    QProgressBar, QProgressDialog
)
from PyQt5.QtGui import QIcon, QFont, QClipboard, QColor, QBrush, QPalette, QIntValidator
from PyQt5.QtCore import Qt, pyqtSignal, QObject, QTimer, QEvent, QSettings, QPoint, QPropertyAnimation, QEasingCurve, QThread

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

import pyzbar.pyzbar as pyzbar
import pyperclip


class UndoManager:
    def __init__(self, max_size=10):
        self.undo_stack = deque(maxlen=max_size)
        self.redo_stack = deque(maxlen=max_size)
        
    def add_action(self, action):
        self.undo_stack.append(action)
        self.redo_stack.clear()
        
    def can_undo(self):
        return len(self.undo_stack) > 0
        
    def can_redo(self):
        return len(self.redo_stack) > 0
        
    def undo(self):
        if not self.undo_stack:
            return None
        action = self.undo_stack.pop()
        self.redo_stack.append(action)
        return action
        
    def redo(self):
        if not self.redo_stack:
            return None
        action = self.redo_stack.pop()
        self.undo_stack.append(action)
        return action


class ToolTip(QObject):
    def __init__(self, widget):
        super().__init__()
        self.widget = widget
        self.tipwindow = None
        self._tooltip_text = ""
        self.widget.installEventFilter(self)

    def showtip(self, text):
        if self.tipwindow or not text:
            return

        rect = self.widget.rect()
        point = self.widget.mapToGlobal(rect.bottomLeft())
        x = point.x() + 5
        y = point.y() + 5

        self.tipwindow = QDialog(self.widget, Qt.FramelessWindowHint | Qt.ToolTip)
        self.tipwindow.setStyleSheet("QDialog {background-color: #ffffe0; border: 1px solid black;}")
        layout = QVBoxLayout(self.tipwindow)
        layout.setContentsMargins(5, 5, 5, 5)
        label = QLabel(text, self.tipwindow)
        label.setFont(QFont("Tahoma", 8))
        layout.addWidget(label)
        self.tipwindow.adjustSize()
        self.tipwindow.move(x, y)
        self.tipwindow.show()

    def hidetip(self):
        if self.tipwindow:
            self.tipwindow.close()
            self.tipwindow.destroy()
            self.tipwindow = None

    def eventFilter(self, watched, event):
        if watched == self.widget:
            if event.type() == QEvent.Enter:
                QTimer.singleShot(500, lambda: self.showtip(self._tooltip_text))
            elif event.type() == QEvent.Leave:
                self.hidetip()
        return super().eventFilter(watched, event)

    def setToolTip(self, text):
        self._tooltip_text = text


class ScanNotification(QLabel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet("""
            QLabel {
                background-color: #27ae60;
                color: white;
                font-size: 14px;
                font-weight: bold;
                padding: 10px 20px;
                border-radius: 5px;
                border: 2px solid #2ecc71;
            }
        """)
        self.setAlignment(Qt.AlignCenter)
        self.hide()
        
    def show_notification(self, text, is_warning=False):
        if is_warning:
            self.setStyleSheet("""
                QLabel {
                    background-color: #e74c3c;
                    color: white;
                    font-size: 14px;
                    font-weight: bold;
                    padding: 10px 20px;
                    border-radius: 5px;
                    border: 2px solid #c0392b;
                }
            """)
        else:
            self.setStyleSheet("""
                QLabel {
                    background-color: #27ae60;
                    color: white;
                    font-size: 14px;
                    font-weight: bold;
                    padding: 10px 20px;
                    border-radius: 5px;
                    border: 2px solid #2ecc71;
                }
            """)
        
        self.setText(text)
        self.adjustSize()
        
        parent_rect = self.parent().rect()
        x = (parent_rect.width() - self.width()) // 2
        y = 50
        self.move(x, y)
        
        self.show()
        self.raise_()
        
        self.animation = QPropertyAnimation(self, b"windowOpacity")
        self.animation.setDuration(2000)
        self.animation.setStartValue(1.0)
        self.animation.setEndValue(0.0)
        self.animation.setEasingCurve(QEasingCurve.OutCubic)
        self.animation.finished.connect(self.hide)
        self.animation.start()


class LoaderDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("⏳ Загрузка")
        self.setModal(True)
        self.setFixedSize(400, 150)
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        
        self.setStyleSheet("""
            QDialog {
                background-color: white;
                border: 2px solid #3498db;
                border-radius: 10px;
            }
            QLabel {
                font-size: 14px;
                color: #2c3e50;
            }
        """)
        
        self.status_label = QLabel("⏳ Загрузка...")
        self.status_label.setAlignment(Qt.AlignCenter)
        self.status_label.setWordWrap(True)
        layout.addWidget(self.status_label)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 1px solid #ced4da;
                border-radius: 3px;
                text-align: center;
                height: 20px;
            }
            QProgressBar::chunk {
                background-color: #3498db;
                border-radius: 2px;
            }
        """)
        layout.addWidget(self.progress_bar)
        
        self.cancel_button = QPushButton("Отмена")
        self.cancel_button.clicked.connect(self.reject)
        layout.addWidget(self.cancel_button)
        
    def set_status(self, text):
        self.status_label.setText(text)
        QApplication.processEvents()
        
    def set_progress(self, value, maximum=100):
        self.progress_bar.setVisible(True)
        self.progress_bar.setMaximum(maximum)
        self.progress_bar.setValue(value)
        QApplication.processEvents()
        
    def set_indeterminate(self, text="⏳ Загрузка..."):
        self.progress_bar.setVisible(False)
        self.status_label.setText(text)
        QApplication.processEvents()


class SaveFormatDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("💾 Сохранить данные")
        self.setModal(True)
        self.setFixedSize(500, 350)
        
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        
        title_label = QLabel("Выберите формат сохранения:")
        title_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #2c3e50;")
        layout.addWidget(title_label)
        
        self.format_list = QTreeWidget()
        self.format_list.setColumnCount(1)
        self.format_list.setHeaderHidden(True)
        self.format_list.setIndentation(0)
        self.format_list.setStyleSheet("""
            QTreeWidget {
                border: 1px solid #ced4da;
                border-radius: 5px;
                padding: 5px;
            }
            QTreeWidget::item {
                padding: 10px;
                border-bottom: 1px solid #f0f0f0;
            }
            QTreeWidget::item:selected {
                background-color: #e0f7fa;
                color: #2c3e50;
            }
        """)
        
        formats = [
            ("📊 CSV с историей", "Логирование действий, НЕ СОХРАНЯТЬ ОТДЕЛЬНО ОТ EXCEL!"),
            ("📑 Excel (несколько листов)", "Каждый короб на отдельном листе"),
            ("📋 Excel (один лист)", "Все данные на одном листе")
        ]
        
        for title, desc in formats:
            item = QTreeWidgetItem([title])
            item.setToolTip(0, desc)
            self.format_list.addTopLevelItem(item)
        
        self.format_list.setCurrentItem(self.format_list.topLevelItem(0))
        layout.addWidget(self.format_list)
        
        desc_label = QLabel("")
        desc_label.setStyleSheet("color: #7f8c8d; font-style: italic; padding: 5px;")
        layout.addWidget(desc_label)
        
        self.format_list.currentItemChanged.connect(
            lambda: desc_label.setText(self.format_list.currentItem().toolTip(0))
        )
        desc_label.setText(self.format_list.topLevelItem(0).toolTip(0))
        
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        
        ok_button = buttons.button(QDialogButtonBox.Ok)
        ok_button.setText("Сохранить")
        ok_button.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                font-weight: bold;
                padding: 8px 20px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        
        cancel_button = buttons.button(QDialogButtonBox.Cancel)
        cancel_button.setText("Отмена")
        
        layout.addWidget(buttons)
        
    def get_selected_format(self):
        current = self.format_list.currentItem()
        if current:
            index = self.format_list.indexOfTopLevelItem(current)
            return index
        return 0


class LoaderThread(QThread):
    progress_update = pyqtSignal(int, int)
    status_update = pyqtSignal(str)
    finished_loading = pyqtSignal(object)
    error_occurred = pyqtSignal(str)
    
    def __init__(self, func, *args, **kwargs):
        super().__init__()
        self.func = func
        self.args = args
        self.kwargs = kwargs
        
    def run(self):
        try:
            def progress_callback(value, maximum):
                self.progress_update.emit(value, maximum)
                
            def status_callback(text):
                self.status_update.emit(text)
            
            self.kwargs['progress_callback'] = progress_callback
            self.kwargs['status_callback'] = status_callback
            
            result = self.func(*self.args, **self.kwargs)
            self.finished_loading.emit(result)
        except Exception as e:
            self.error_occurred.emit(str(e))


class InvoiceViewDialog(QDialog):
    def __init__(self, invoice_data, filename, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"📋 Накладная: {filename}")
        self.setGeometry(200, 200, 600, 500)
        self.setModal(True)
        
        layout = QVBoxLayout(self)
        
        total_items = len(invoice_data)
        total_quantity = sum(invoice_data.values())
        info_label = QLabel(f"Позиций: {total_items} | Всего товаров: {total_quantity} шт")
        info_label.setStyleSheet("font-weight: bold; color: #3498db;")
        layout.addWidget(info_label)
        
        self.table = QTableWidget()
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(["Штрихкод", "Количество"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        
        self.table.setRowCount(len(invoice_data))
        for row, (barcode, count) in enumerate(invoice_data.items()):
            self.table.setItem(row, 0, QTableWidgetItem(barcode))
            count_item = QTableWidgetItem(str(count))
            count_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 1, count_item)
        
        layout.addWidget(self.table)
        
        close_button = QPushButton("Закрыть")
        close_button.clicked.connect(self.accept)
        layout.addWidget(close_button)


class ReportDialog(QDialog):
    def __init__(self, report_text, parent=None):
        super().__init__(parent)
        self.setWindowTitle("📄 Отчёт о сборке")
        self.setGeometry(200, 200, 800, 600)
        self.setModal(True)
        
        layout = QVBoxLayout(self)
        
        self.text_edit = QTextEdit()
        self.text_edit.setReadOnly(True)
        self.text_edit.setFont(QFont("Courier New", 10))
        self.text_edit.setText(report_text)
        layout.addWidget(self.text_edit)
        
        buttons = QDialogButtonBox(QDialogButtonBox.Ok)
        buttons.accepted.connect(self.accept)
        layout.addWidget(buttons)


class EditCountDialog(QDialog):
    def __init__(self, barcode, current_count, planned=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("✏️ Изменить количество")
        self.setModal(True)
        self.setMinimumWidth(350)
        self.setMaximumHeight(400)
        self.setWindowFlags(Qt.Dialog | Qt.WindowCloseButtonHint | Qt.WindowStaysOnTopHint)
        
        layout = QVBoxLayout(self)
        layout.setSpacing(8)
        layout.setContentsMargins(12, 12, 12, 12)
        
        # Информация в одну строку (компактно)
        info_widget = QWidget()
        info_layout = QHBoxLayout(info_widget)
        info_layout.setContentsMargins(0, 0, 0, 0)
        info_layout.setSpacing(10)
        
        # Штрихкод
        barcode_container = QWidget()
        barcode_container.setStyleSheet("background-color: #f0f0f0; border-radius: 4px;")
        barcode_layout = QHBoxLayout(barcode_container)
        barcode_layout.setContentsMargins(8, 4, 8, 4)
        
        barcode_icon = QLabel("📦")
        barcode_layout.addWidget(barcode_icon)
        
        self.barcode_value = QLabel(barcode)
        self.barcode_value.setStyleSheet("font-family: 'Courier New'; font-weight: bold;")
        barcode_layout.addWidget(self.barcode_value)
        
        info_layout.addWidget(barcode_container)
        
        # Текущее количество
        current_container = QWidget()
        current_container.setStyleSheet("background-color: #e8f5e9; border-radius: 4px;")
        current_layout = QHBoxLayout(current_container)
        current_layout.setContentsMargins(8, 4, 8, 4)
        
        current_icon = QLabel("📊")
        current_layout.addWidget(current_icon)
        
        self.current_value = QLabel(str(current_count))
        self.current_value.setStyleSheet("font-weight: bold; color: #2e7d32;")
        current_layout.addWidget(self.current_value)
        
        info_layout.addWidget(current_container)
        
        # План (если есть)
        if planned is not None:
            planned_container = QWidget()
            planned_container.setStyleSheet("background-color: #fff3e0; border-radius: 4px;")
            planned_layout = QHBoxLayout(planned_container)
            planned_layout.setContentsMargins(8, 4, 8, 4)
            
            planned_icon = QLabel("📋")
            planned_layout.addWidget(planned_icon)
            
            self.planned_value = QLabel(str(planned))
            self.planned_value.setStyleSheet("font-weight: bold; color: #e65100;")
            planned_layout.addWidget(self.planned_value)
            
            info_layout.addWidget(planned_container)
        
        layout.addWidget(info_widget)
        
        # Поле ввода (крупное, чтобы было удобно тыкать)
        self.count_edit = QLineEdit(str(current_count))
        self.count_edit.setAlignment(Qt.AlignCenter)
        self.count_edit.setMinimumHeight(60)
        self.count_edit.setStyleSheet("""
            QLineEdit {
                font-size: 24pt;
                font-weight: bold;
                border: 2px solid #2196f3;
                border-radius: 6px;
                margin: 5px 0;
            }
            QLineEdit:focus {
                border-color: #ff9800;
            }
        """)
        layout.addWidget(self.count_edit)
        
        # Кнопки +/- крупные
        steps_widget = QWidget()
        steps_layout = QHBoxLayout(steps_widget)
        steps_layout.setContentsMargins(0, 0, 0, 0)
        steps_layout.setSpacing(5)
        
        self.minus_button = QPushButton("−")
        self.minus_button.setFixedHeight(50)
        self.minus_button.setStyleSheet("""
            QPushButton {
                font-size: 20pt;
                background-color: #ffebee;
                color: #c62828;
                border: none;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #ffcdd2;
            }
        """)
        self.minus_button.clicked.connect(self.decrement_value)
        steps_layout.addWidget(self.minus_button)
        
        self.plus_button = QPushButton("+")
        self.plus_button.setFixedHeight(50)
        self.plus_button.setStyleSheet("""
            QPushButton {
                font-size: 20pt;
                background-color: #e8f5e9;
                color: #2e7d32;
                border: none;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #c8e6c9;
            }
        """)
        self.plus_button.clicked.connect(self.increment_value)
        steps_layout.addWidget(self.plus_button)
        
        layout.addWidget(steps_widget)
        
        # Кнопки быстрого изменения (в виде полосы)
        quick_widget = QWidget()
        quick_layout = QHBoxLayout(quick_widget)
        quick_layout.setContentsMargins(0, 0, 0, 0)
        quick_layout.setSpacing(3)
        
        for value in [1, 5, 10, 50, 100]:
            btn = QPushButton(f"{value:+d}")
            btn.setFixedHeight(35)
            btn.setStyleSheet("""
                QPushButton {
                    background-color: #e3f2fd;
                    color: #1565c0;
                    border: none;
                    border-radius: 3px;
                    font-weight: bold;
                    font-size: 9pt;
                }
                QPushButton:hover {
                    background-color: #bbdefb;
                }
            """)
            btn.clicked.connect(lambda checked, v=value: self.add_value(v))
            quick_layout.addWidget(btn)
        
        layout.addWidget(quick_widget)
        
        # Кнопки действий
        buttons_widget = QWidget()
        buttons_layout = QHBoxLayout(buttons_widget)
        buttons_layout.setContentsMargins(0, 5, 0, 0)
        buttons_layout.setSpacing(5)
        
        self.ok_button = QPushButton("✅ Применить")
        self.ok_button.setFixedHeight(40)
        self.ok_button.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                font-weight: bold;
                border: none;
                border-radius: 4px;
                font-size: 11pt;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)
        self.ok_button.clicked.connect(self.accept)
        buttons_layout.addWidget(self.ok_button)
        
        self.cancel_button = QPushButton("✕ Отмена")
        self.cancel_button.setFixedHeight(40)
        self.cancel_button.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                font-weight: bold;
                border: none;
                border-radius: 4px;
                font-size: 11pt;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        self.cancel_button.clicked.connect(self.reject)
        buttons_layout.addWidget(self.cancel_button)
        
        layout.addWidget(buttons_widget)
        
        # Настройки ввода
        self.count_edit.setFocus()
        self.count_edit.selectAll()
        self.count_edit.setValidator(QIntValidator(0, 999999))
        self.count_edit.returnPressed.connect(self.accept)
        
    def increment_value(self):
        try:
            current = int(self.count_edit.text())
            self.count_edit.setText(str(current + 1))
        except:
            self.count_edit.setText("1")
            
    def decrement_value(self):
        try:
            current = int(self.count_edit.text())
            if current > 0:
                self.count_edit.setText(str(current - 1))
        except:
            pass
            
    def add_value(self, delta):
        try:
            current = int(self.count_edit.text())
            new_value = max(0, current + delta)
            self.count_edit.setText(str(new_value))
        except:
            self.count_edit.setText(str(max(0, delta)))
    
    def get_value(self):
        try:
            return int(self.count_edit.text())
        except:
            return 0


class ConfirmationDialog(QDialog):
    def __init__(self, title, message, icon_type="question", parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setModal(True)
        self.setFixedSize(400, 200)
        self.setWindowFlags(Qt.Dialog | Qt.WindowCloseButtonHint)
        
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Иконка и сообщение
        content_widget = QWidget()
        content_layout = QHBoxLayout(content_widget)
        content_layout.setSpacing(15)
        
        # Выбор иконки
        if icon_type == "question":
            icon_text = "❓"
            icon_color = "#3b82f6"
        elif icon_type == "warning":
            icon_text = "⚠️"
            icon_color = "#f59e0b"
        elif icon_type == "error":
            icon_text = "❌"
            icon_color = "#ef4444"
        else:  # info
            icon_text = "ℹ️"
            icon_color = "#3b82f6"
        
        icon_label = QLabel(icon_text)
        icon_label.setStyleSheet(f"font-size: 48px; color: {icon_color};")
        icon_label.setAlignment(Qt.AlignCenter)
        icon_label.setFixedSize(80, 80)
        content_layout.addWidget(icon_label)
        
        message_label = QLabel(message)
        message_label.setStyleSheet("font-size: 14px; color: #1e293b;")
        message_label.setWordWrap(True)
        content_layout.addWidget(message_label)
        
        layout.addWidget(content_widget)
        
        layout.addStretch()
        
        # Кнопки
        buttons_widget = QWidget()
        buttons_layout = QHBoxLayout(buttons_widget)
        buttons_layout.setContentsMargins(0, 0, 0, 0)
        buttons_layout.setSpacing(10)
        
        buttons_layout.addStretch()
        
        self.yes_button = QPushButton("✅ Да")
        self.yes_button.setFixedSize(100, 40)
        self.yes_button.setStyleSheet("""
            QPushButton {
                background-color: #3b82f6;
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #2563eb;
            }
        """)
        self.yes_button.clicked.connect(self.accept)
        buttons_layout.addWidget(self.yes_button)
        
        self.no_button = QPushButton("✕ Нет")
        self.no_button.setFixedSize(100, 40)
        self.no_button.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        self.no_button.clicked.connect(self.reject)
        buttons_layout.addWidget(self.no_button)
        
        layout.addWidget(buttons_widget)


class QBarcodeApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("ScanBox")

        if getattr(sys, '_MEIPASS', None):
            base_path = sys._MEIPASS
        else:
            base_path = Path(__file__).resolve().parent

        self.scan_notification = None
        self.log_dir = os.path.join(base_path, "logs")
        os.makedirs(self.log_dir, exist_ok=True)

        self.all_boxes = {}
        self.current_box_barcode = ""
        self.search_query = ""
        self.packer_name = ""
        
        self.invoice_data = {}
        self.invoice_loaded = False
        self.invoice_file_name = ""
        self.invoice_file_path = ""
        
        self.scan_history = []
        self.undo_manager = UndoManager(max_size=10)
        
        self.start_time = None
        self.paused_time = None
        self.pause_start = None
        self.is_paused = False
        self.total_scans = 0
        self.has_unsaved_changes = False
        self.first_scan_done = False
        
        self.update_timer = QTimer()
        self.update_timer.timeout.connect(self.update_stats)
        self.update_timer.start(1000)
        
        self.state_file_dir = Path(os.path.expanduser("~")) / ".ScanBox"
        os.makedirs(self.state_file_dir, exist_ok=True)
        self.state_file = str(self.state_file_dir / "barcode_app_state.json")
        
        self.settings = QSettings("ScanBox", "ScanBox")

        self.history_window = None
        self.history_tree = None
        self.comments = {}
        self.history_filter_query = ""

        self.COLOR_BG = "#f8f9fa"
        self.COLOR_FRAME_BG = "#ffffff"
        self.COLOR_ENTRY_BG = "#ffffff"
        self.COLOR_BUTTON_BG = "#e0f7fa"
        self.COLOR_BUTTON_FG = "#495057"
        self.COLOR_BUTTON_ACTIVE_BG = "#b2ebf2"
        self.COLOR_BUTTON_PRESSED_BG = "#80deea"
        self.COLOR_HEADER_BG = "#f8f9fa"
        self.COLOR_HEADER_FG = "#212529"
        self.COLOR_SCROLLBAR_BG = "#f1f3f5"
        self.COLOR_SCROLLBAR_TROUGH = "#ced4da"
        self.COLOR_SCROLLBAR_ARROW = "black"
        self.COLOR_SCROLLBAR_THUMB = "#949ca1"

        self.font_label = QFont("Segoe UI", 9)
        self.font_entry = QFont("Segoe UI", 9)
        self.font_button = QFont("Segoe UI Semibold", 10)
        self.font_treeview = QFont("Segoe UI", 9)
        self.font_treeview_heading = QFont("Segoe UI Semibold", 11)
        self.font_menu = QFont("Segoe UI", 10)

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.main_layout = QVBoxLayout()
        self.central_widget.setLayout(self.main_layout)
        self.main_layout.setContentsMargins(10, 10, 10, 10)

        self.strict_validation_enabled = True
        self.loader_dialog = None
        self.loader_thread = None

        self.create_top_info_frame()
        self.create_menu_bar()
        self.create_search_frame()
        self.create_box_frame()
        self.create_item_scan_frame()
        self.create_items_frame()
        self.create_stats_frame()
        self.create_control_frame()
        self.create_status_bar()

        self.load_state()
        self.load_column_settings()

        self.clipboard = QApplication.clipboard()

        self.setAcceptDrops(True)
        self.drop_indicator = None
        self.create_drop_indicator()
        
        self.scan_notification = ScanNotification(self.central_widget)

        self.setStyleSheet(self.get_stylesheet())

        self.setGeometry(100, 100, 1280, 720)

    def show_loader(self, func, *args, **kwargs):
        self.loader_dialog = LoaderDialog(self)
        self.loader_thread = LoaderThread(func, *args, **kwargs)
    
        self.loader_thread.progress_update.connect(
            lambda v, m: self.loader_dialog.set_progress(v, m)
        )
        self.loader_thread.status_update.connect(
            lambda t: self.loader_dialog.set_status(t)
        )
        self.loader_thread.finished_loading.connect(self.on_loader_finished)
        self.loader_thread.error_occurred.connect(self.on_loader_error)
    
        self.loader_dialog.rejected.connect(self.on_loader_cancelled)
        
        self.loader_dialog.show()
        self.loader_thread.start()

    def on_loader_cancelled(self):
        if self.loader_thread and self.loader_thread.isRunning():
            self.loader_thread.terminate()
            self.loader_thread.wait()
        self.loader_thread = None
        self.loader_dialog = None
        self.update_status("Загрузка отменена")
        
    def on_loader_finished(self, result):
        if self.loader_dialog:
            self.loader_dialog.accept()
            self.loader_dialog = None
            
        if isinstance(result, tuple):
            if len(result) == 5:  # Excel результат
                invoice_data, total_items, total_quantity, file_name, file_path = result
                
                self.invoice_data = invoice_data
                self.invoice_loaded = True
                self.invoice_file_name = file_name
                self.invoice_file_path = file_path
                self.invoice_label.setText(f"📋 Накладная: {file_name} (позиций: {total_items}, всего: {total_quantity} шт)")
                self.clear_invoice_button.setEnabled(True)
                self.view_invoice_button.setEnabled(True)
                self.pause_button.show()
                
                self.start_time = None
                self.first_scan_done = False
                self.is_paused = False
                self.pause_button.setText("⏸️")
                
                self.refresh_treeview()
                self.update_status(f"✅ Загружена накладная: {file_name}")
                self.status_bar.showMessage(f"✅ Накладная загружена: {file_name}", 5000)
                
                QMessageBox.information(self, "Успешно", f"Загружено {total_items} позиций, всего {total_quantity} шт")
                
            elif len(result) == 7:  # CSV результат
                all_boxes, comments, scan_history, packer_name, start_time, first_scan_done, file_name = result
                
                self.all_boxes = all_boxes
                self.comments = comments
                self.scan_history = scan_history
                
                # Создаем новый UndoManager
                self.undo_manager = UndoManager(max_size=10)
                
                if packer_name:
                    self.packer_name = packer_name
                    self.packer_combo.setCurrentText(packer_name)
                self.start_time = start_time
                self.first_scan_done = first_scan_done
                self.has_unsaved_changes = False
                
                self.refresh_treeview()
                if self.all_boxes:
                    self.update_status(f"✅ Данные загружены из {file_name}")
                    self.save_button.setEnabled(True)
                    self.status_bar.showMessage(f"✅ Файл {file_name} успешно загружен!", 5000)
                    
                    if self.history_window and self.history_window.isVisible():
                        self.populate_history_tree()
        
        self.loader_thread = None
        
    def on_loader_error(self, error_msg):
        if self.loader_dialog:
            self.loader_dialog.accept()
            self.loader_dialog = None
        self.loader_thread = None
        self.show_error(f"Ошибка загрузки: {error_msg}")

    def create_top_info_frame(self):
        self.top_frame = QGroupBox("👤 Информация о сборщике")
        self.main_layout.addWidget(self.top_frame)
        top_layout = QHBoxLayout()
        self.top_frame.setLayout(top_layout)

        self.packer_label = QLabel("Сборщик:")
        self.packer_label.setStyleSheet("""
            QLabel {
                font-size: 10pt;
                font-weight: bold;
                color: #2c3e50;
            }
        """)
        top_layout.addWidget(self.packer_label)

        self.packer_combo = QComboBox()
        self.packer_combo.setEditable(True)
        self.packer_combo.setMinimumWidth(200)
        self.packer_combo.setMaximumWidth(250)
        self.packer_combo.setMinimumHeight(28)
        self.packer_combo.setMaximumHeight(30)
    
        self.packer_combo.setStyleSheet("""
            QComboBox {
                background-color: white;
                border: 1px solid #3498db;
                border-radius: 4px;
                padding: 3px 8px;
                font-size: 10pt;
                color: #2c3e50;
            }
            QComboBox:hover {
                border-color: #2980b9;
                background-color: #f0f8ff;
            }
            QComboBox:focus {
                border-color: #e67e22;
            }
            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 22px;
                border-left: 1px solid #3498db;
                border-top-right-radius: 3px;
                border-bottom-right-radius: 3px;
                background-color: #e0f7fa;
            }
            QComboBox::down-arrow {
                width: 10px;
                height: 10px;
                image: none;
                border-left: 4px solid transparent;
                border-right: 4px solid transparent;
                border-top: 6px solid #2c3e50;
                margin-top: 2px;
            }
            QComboBox::drop-down:hover {
                background-color: #b2ebf2;
            }
            QComboBox QAbstractItemView {
                background-color: white;
                border: 1px solid #3498db;
                border-radius: 4px;
                padding: 3px;
                outline: none;
                selection-background-color: #e0f7fa;
                selection-color: #2c3e50;
                font-size: 10pt;
            }
            QComboBox QAbstractItemView::item {
                padding: 5px 8px;
                border-bottom: 1px solid #ecf0f1;
                color: #2c3e50;
            }
            QComboBox QAbstractItemView::item:selected {
                background-color: #3498db;
                color: white;
            }
            QComboBox QAbstractItemView::item:hover {
                background-color: #e0f7fa;
                color: #2c3e50;
            }
        """)
    
        self.packer_combo.addItem("👤 Не выбран", "")
        self.packer_combo.addItem("👨‍💻 Алексей Корябкин", "Алексей Корябкин")
        self.packer_combo.addItem("👨‍💻 Илья Богданов", "Илья Богданов")
        self.packer_combo.addItem("👨‍💻 Никита Овчинников", "Никита Овчинников")
        self.packer_combo.addItem("👥 РГ", "РГ")
        self.packer_combo.addItem("👥 Несколько сотрудников", "Несколько сотрудников")
    
        self.packer_combo.currentIndexChanged.connect(self.on_packer_changed)
        top_layout.addWidget(self.packer_combo)

        self.invoice_label = QLabel("")
        self.invoice_label.setStyleSheet("color: #27ae60; font-weight: bold; margin-left: 15px; font-size: 9pt;")
        top_layout.addWidget(self.invoice_label)
    
        top_layout.addStretch()

    def on_packer_changed(self, index):
        if index >= 0:
            self.packer_name = self.packer_combo.itemData(index)
            if self.packer_name is None:
                self.packer_name = self.packer_combo.currentText().strip()
        else:
            self.packer_name = self.packer_combo.currentText().strip()
        self.save_state()

    def create_drop_indicator(self):
        self.drop_indicator = QLabel(self.centralWidget())
        self.drop_indicator.setAlignment(Qt.AlignCenter)
        self.drop_indicator.setStyleSheet("""
            QLabel {
                background-color: rgba(52, 152, 219, 0.2);
                border: 4px dashed #3498db;
                border-radius: 10px;
                font-size: 24px;
                font-weight: bold;
                color: #2980b9;
            }
        """)
        self.drop_indicator.setText("📄 Перетащите CSV или Excel файл сюда\nдля импорта данных")
        self.drop_indicator.setGeometry(0, 0, self.width(), self.height())
        self.drop_indicator.hide()
        
        # Анимация появления/исчезновения
        self.drop_animation = QPropertyAnimation(self.drop_indicator, b"windowOpacity")
        self.drop_animation.setDuration(200)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            for url in event.mimeData().urls():
                file_path = url.toLocalFile()
                if file_path.lower().endswith(('.csv', '.xlsx', '.xls')):
                    if self.drop_indicator:
                        self.drop_indicator.show()
                        self.drop_indicator.raise_()
                    event.acceptProposedAction()
                    return
        event.ignore()

    def dragLeaveEvent(self, event):
        if self.drop_indicator:
            self.drop_indicator.hide()
        event.accept()

    def dropEvent(self, event):
        if self.drop_indicator:
            self.drop_indicator.hide()
    
        files = []
        for url in event.mimeData().urls():
            file_path = url.toLocalFile()
            if file_path.lower().endswith(('.csv', '.xlsx', '.xls')):
                files.append(file_path)
    
        if files:
            if len(files) > 1:
                self.show_multiple_files_dialog(files)
            else:
                self.import_file(files[0])
            event.acceptProposedAction()

    def show_multiple_files_dialog(self, files):
        msg = QMessageBox(self)
        msg.setWindowTitle("Несколько файлов")
        msg.setText("Вы перетащили несколько файлов. Выберите действие:")
        
        for file_path in files:
            btn = msg.addButton(f"Загрузить: {os.path.basename(file_path)}", QMessageBox.ActionRole)
            btn.clicked.connect(lambda checked, path=file_path: self.import_file(path))
        
        msg.addButton(QMessageBox.Cancel)
        msg.exec_()

    def import_file(self, file_path):
        self.show_loader(self._import_file_task, file_path)
        
    def _import_file_task(self, file_path, progress_callback=None, status_callback=None):
        if status_callback:
            status_callback(f"📂 Загрузка {os.path.basename(file_path)}...")
            
        ext = os.path.splitext(file_path)[1].lower()
        
        if ext == '.csv':
            self._drag_import_file = file_path
            result = self._load_csv_task(file_path, progress_callback, status_callback)
        elif ext in ('.xlsx', '.xls'):
            result = self._load_invoice_task(file_path, progress_callback, status_callback)
        else:
            raise Exception("Неподдерживаемый формат файла")
        
        return result

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if self.drop_indicator:
            self.drop_indicator.setGeometry(0, 0, self.width(), self.height())

    def get_stylesheet(self):
        return f"""
            QMainWindow {{
                background-color: {self.COLOR_BG};
            }}
            QFrame {{
                background-color: {self.COLOR_FRAME_BG};
            }}
            QGroupBox {{
                background-color: {self.COLOR_FRAME_BG};
                border: 1px groove #ced4da;
                border-radius: 2px;
                margin-top: 0.5em;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 3px 0 3px;
                font: bold 11pt "Segoe UI Semibold";
                color: {self.COLOR_HEADER_FG};
                background-color: {self.COLOR_FRAME_BG};
            }}
            QLabel {{
                font: 9pt "Segoe UI";
                background-color: {self.COLOR_BG};
            }}
            QLineEdit {{
                font: 9pt "Segoe UI";
                background-color: {self.COLOR_ENTRY_BG};
                border: 1px solid #ced4da;
                border-radius: 2px;
                padding: 2px;
            }}
            QPushButton {{
                font: bold 10pt "Segoe UI Semibold";
                background-color: {self.COLOR_BUTTON_BG};
                color: {self.COLOR_BUTTON_FG};
                border: none;
                padding: 8px 18px;
                border-radius: 6px;
                box-shadow: 2px 2px 3px rgba(0,0,0,0.1);
            }}
            QPushButton:hover {{
                background-color: {self.COLOR_BUTTON_ACTIVE_BG};
                box-shadow: 3px 3px 5px rgba(0,0,0,0.15);
            }}
            QPushButton:pressed {{
                background-color: {self.COLOR_BUTTON_PRESSED_BG};
                box-shadow: 1px 1px 2px rgba(0,0,0,0.1);
            }}
            QPushButton#undo_button {{
                background-color: #fff3e0;
                color: #e67e22;
            }}
            QPushButton#undo_button:hover {{
                background-color: #ffe0b2;
            }}
            QPushButton#undo_button:disabled {{
                background-color: #f5f5f5;
                color: #bdbdbd;
            }}
            QMenuBar {{
                background-color: {self.COLOR_HEADER_BG};
                color: {self.COLOR_HEADER_FG};
                border-bottom: 1px solid #ced4da;
            }}
            QMenuBar::item {{
                background-color: transparent;
            }}
            QMenuBar::item:selected {{
                background-color: #e0e0e0;
            }}
            QMenu {{
                font: 10pt "Segoe UI";
                background-color: {self.COLOR_FRAME_BG};
                border: 1px solid #ced4da;
            }}
            QMenu::item:selected {{
                background-color: #bbdefb;
            }}
            QTreeWidget {{
                font: 9pt "Segoe UI";
                background-color: white;
                alternate-background-color: #f0f0f0;
                border: 1px solid #ced4da;
            }}
            QHeaderView::section {{
                background-color: {self.COLOR_HEADER_BG};
                font: bold 11pt "Segoe UI Semibold";
                border: none;
                border-bottom: 1px solid #ced4da;
                padding: 4px;
                qproperty-alignment: AlignCenter;
            }}
            QTreeWidget::item:selected {{
                background-color: #bbdefb;
                color: black;
            }}
            QScrollBar:vertical {{
                background-color: {self.COLOR_SCROLLBAR_BG};
                width: 10px;
                margin: 0px 0px 0px 0px;
            }}
            QScrollBar::handle:vertical {{
                background-color: {self.COLOR_SCROLLBAR_THUMB};
                min-height: 20px;
                border-radius: 5px;
            }}
            QScrollBar::add-line:vertical {{
                height: 0px;
                subcontrol-position: bottom;
                subcontrol-origin: margin;
            }}
            QScrollBar::sub-line:vertical {{
                height: 0 px;
                subcontrol-position: top left;
                subcontrol-origin: margin;
            }}
            QScrollBar::up-arrow:vertical, QScrollBar::down-arrow:vertical {{
                background: none;
            }}
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{
                background: none;
            }}
            QScrollBar:horizontal {{
                background-color: {self.COLOR_SCROLLBAR_BG};
                height: 10px;
                margin: 0px 0px 0px 0px;
            }}
            QScrollBar::handle:horizontal {{
                background-color: {self.COLOR_SCROLLBAR_THUMB};
                min-width: 20px;
                border-radius: 5px;
            }}
            QScrollBar::add-line:horizontal {{
                width: 0px;
                subcontrol-position: right;
                subcontrol-origin: margin;
            }}
            QScrollBar::sub-line:horizontal {{
                width: 0 px;
                subcontrol-position: left top;
                subcontrol-origin: margin;
            }}
            QScrollBar::left-arrow:horizontal, QScrollBar::right-arrow:horizontal {{
                background: none;
            }}
            QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {{
                background: none;
            }}
            QStatusBar {{
                background-color: {self.COLOR_HEADER_BG};
                color: black;
                border-top: 1px solid #ced4da;
            }}
            QCheckBox {{
                font: 9pt "Segoe UI";
                background-color: {self.COLOR_FRAME_BG};
            }}
            QTextEdit {{
                font: 9pt "Segoe UI";
                background-color: white;
                border: 1px solid #ced4da;
            }}
            QComboBox {{
                font: 9pt "Segoe UI";
                background-color: {self.COLOR_ENTRY_BG};
                border: 1px solid #ced4da;
                border-radius: 2px;
                padding: 2px;
                min-height: 20px;
                padding-right: 25px;
            }}
            QComboBox:hover {{
                border-color: #3498db;
            }}
            QComboBox::drop-down {{
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 25px;
                border-left-width: 1px;
                border-left-color: #3498db;
                border-left-style: solid;
                background-color: #e0f7fa;
            }}
            QComboBox::down-arrow {{
                width: 12px;
                height: 12px;
            }}
            QComboBox::drop-down:hover {{
                background-color: #b2ebf2;
            }}
            QHeaderView::section {{
                background-color: #e9ecef;
                padding: 8px;
                border: 1px solid #dee2e6;
                font-weight: bold;
            }}
            QTableWidget {{
                font: 9pt "Segoe UI";
                background-color: white;
                alternate-background-color: #f8f9fa;
            }}
            QProgressBar {{
                border: 1px solid #ced4da;
                border-radius: 3px;
                text-align: center;
                background-color: #f8f9fa;
            }}
            QProgressBar::chunk {{
                background-color: #3498db;
                border-radius: 2px;
            }}
        """

    def create_menu_bar(self):
        menubar = QMenuBar(self)
        self.setMenuBar(menubar)

        menu_menu = menubar.addMenu("Меню")

        action_reset = QAction("🔄 Начать заново", self)
        action_reset.setShortcut("Ctrl+R")
        action_reset.triggered.connect(self.reset_application)
        menu_menu.addAction(action_reset)

        menu_menu.addSeparator()

        action_save = QAction("💾 Сохранить...", self)
        action_save.setShortcut("Ctrl+S")
        action_save.triggered.connect(self.save_with_format_dialog)
        menu_menu.addAction(action_save)

        action_report = QAction("📄 Экспорт отчёта...", self)
        action_report.setShortcut("Ctrl+Shift+R")
        action_report.triggered.connect(self.export_report)
        menu_menu.addAction(action_report)

        import_export_menu = menu_menu.addMenu("📁 Импорт/Экспорт")

        action_load_csv = QAction("📂 Загрузить CSV...", import_export_menu)
        action_load_csv.triggered.connect(self.load_from_csv_dialog)
        import_export_menu.addAction(action_load_csv)
        
        action_load_invoice = QAction("📋 Загрузить накладную Excel...", import_export_menu)
        action_load_invoice.triggered.connect(self.load_invoice_dialog)
        import_export_menu.addAction(action_load_invoice)
        
        action_clear_invoice = QAction("❌ Сбросить накладную", import_export_menu)
        action_clear_invoice.triggered.connect(self.clear_invoice)
        import_export_menu.addAction(action_clear_invoice)

        menu_menu.addSeparator()

        action_settings = QAction("⚙️ Настройки...", menu_menu)
        action_settings.triggered.connect(self.show_settings_dialog)
        menu_menu.addAction(action_settings)
        menu_menu.addSeparator()

        action_exit = QAction("🚪 Закрыть", self)
        action_exit.setShortcut("Ctrl+Q")
        action_exit.triggered.connect(self.close)
        menu_menu.addAction(action_exit)

    def closeEvent(self, event):
        if self.has_unsaved_changes:
            dialog = ConfirmationDialog(
                "⚠️ Несохранённые изменения",
                "У вас есть несохранённые изменения. Сохранить перед выходом?",
                "warning",
                self
            )
            
            # Добавляем третью кнопку "Отмена"
            dialog.yes_button.setText("✅ Да")
            dialog.no_button.setText("✕ Нет")
            
            # Создаем кнопку отмены
            cancel_button = QPushButton("◀ Отмена")
            cancel_button.setFixedSize(100, 40)
            cancel_button.setStyleSheet("""
                QPushButton {
                    background-color: #94a3b8;
                    color: white;
                    border: none;
                    border-radius: 8px;
                    font-weight: bold;
                    font-size: 12px;
                }
                QPushButton:hover {
                    background-color: #64748b;
                }
            """)
            cancel_button.clicked.connect(dialog.reject)
            
            # Добавляем кнопки в правильном порядке
            layout = dialog.layout()
            buttons_widget = layout.itemAt(layout.count() - 1).widget()
            buttons_layout = buttons_widget.layout()
            buttons_layout.insertWidget(0, cancel_button)
            
            reply = dialog.exec_()
            
            if reply == QDialog.Accepted:
                self.save_with_format_dialog()
                event.accept()
            elif reply == QDialog.Rejected:
                # Определяем какая кнопка была нажата
                sender = self.sender()
                if sender == cancel_button:
                    event.ignore()
                else:
                    event.accept()
            else:
                event.ignore()

    def export_report(self):
        if not self.all_boxes:
            self.show_warning("Нет данных для отчёта!")
            return
            
        report_lines = []
        report_lines.append("=" * 80)
        report_lines.append("ОТЧЁТ О СБОРКЕ".center(80))
        report_lines.append("=" * 80)
        report_lines.append(f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
        report_lines.append(f"Сборщик: {self.packer_name if self.packer_name else 'не указан'}")
        if self.invoice_loaded:
            total_planned = sum(self.invoice_data.values())
            report_lines.append(f"Накладная: {self.invoice_file_name} (позиций: {len(self.invoice_data)}, всего: {total_planned} шт)")
        report_lines.append("")
        
        total_scanned = 0
        match_count = 0
        shortage_count = 0
        excess_count = 0
        extra_count = 0
        
        shortage_units = 0
        excess_units = 0
        extra_units = 0
        
        for box_barcode, items in self.all_boxes.items():
            box_comment = self.comments.get((box_barcode, ""), "")
            report_lines.append(f"\n📦 КОРОБ: {box_barcode}")
            if box_comment:
                report_lines.append(f"   Комментарий: {box_comment}")
            
            for item_barcode, count in items.items():
                item_comment = self.comments.get((box_barcode, item_barcode), "")
                total_scanned += count
                
                status = ""
                if self.invoice_loaded:
                    if item_barcode in self.invoice_data:
                        planned = self.invoice_data[item_barcode]
                        total_for_item = self.get_total_scanned_for_item(item_barcode)
                        if total_for_item == planned:
                            status = "✅ СОВПАДАЕТ"
                            match_count += 1
                        elif total_for_item < planned:
                            diff = planned - total_for_item
                            status = f"⚠️ НЕДОБОР (план: {planned}, всего: {total_for_item}, не хватает: {diff})"
                            shortage_count += 1
                            shortage_units += diff
                        else:
                            diff = total_for_item - planned
                            status = f"❗ ПЕРЕБОР (план: {planned}, всего: {total_for_item}, лишних: {diff})"
                            excess_count += 1
                            excess_units += diff
                    else:
                        status = "❓ ЛИШНИЙ"
                        extra_count += 1
                        extra_units += count
                
                line = f"   • {item_barcode} - {count} шт."
                if status:
                    line += f" [{status}]"
                if item_comment:
                    line += f" ({item_comment})"
                report_lines.append(line)
        
        report_lines.append("")
        report_lines.append("-" * 80)
        report_lines.append("ИТОГИ:")
        report_lines.append(f"📦 Коробов: {len(self.all_boxes)}")
        report_lines.append(f"📦 Всего товаров: {total_scanned} шт.")
        
        if self.invoice_loaded:
            total_planned = sum(self.invoice_data.values())
            report_lines.append(f"📋 План: {total_planned} шт.")
            report_lines.append(f"✅ Совпадает: {match_count} позиций")
            report_lines.append(f"⚠️ Недобор: {shortage_count} позиций (всего -{shortage_units} шт)")
            report_lines.append(f"❗ Перебор: {excess_count} позиций (всего +{excess_units} шт)")
            report_lines.append(f"❓ Лишние: {extra_count} позиций (всего +{extra_units} шт)")
        
        report_lines.append("=" * 80)
        
        report_text = "\n".join(report_lines)
        
        dialog = ReportDialog(report_text, self)
        dialog.exec_()
        
        dialog = ConfirmationDialog(
            "💾 Сохранить отчёт",
            "Сохранить отчёт в текстовый файл?",
            "question",
            self
        )
        dialog.yes_button.setText("✅ Да")
        dialog.no_button.setText("✕ Нет")
    
        if dialog.exec_() == QDialog.Accepted:
            file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить отчёт", "", "Text Files (*.txt);;All Files (*)")
            if file_path:
                try:
                    with open(file_path, "w", encoding="utf-8") as f:
                        f.write(report_text)
                    self.show_info(f"✅ Отчёт сохранён: {os.path.basename(file_path)}")
                except Exception as e:
                    self.show_error(f"Ошибка при сохранении отчёта: {e}")

    def show_settings_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("⚙️ Настройки")
        dialog.setModal(True)
        dialog.setFixedSize(450, 250)
        dialog.setWindowFlags(Qt.Dialog | Qt.WindowCloseButtonHint)
        
        layout = QVBoxLayout(dialog)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Заголовок
        title_label = QLabel("Настройки приложения")
        title_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #1e293b; margin-bottom: 10px;")
        layout.addWidget(title_label)
        
        # Валидация штрихкода
        validation_widget = QWidget()
        validation_layout = QHBoxLayout(validation_widget)
        validation_layout.setContentsMargins(0, 0, 0, 0)
        
        validation_icon = QLabel("🔍")
        validation_icon.setStyleSheet("font-size: 20px;")
        validation_layout.addWidget(validation_icon)
        
        validation_text = QLabel("Строгая валидация штрихкода:")
        validation_text.setStyleSheet("font-size: 12px; color: #334155;")
        validation_layout.addWidget(validation_text)
        
        validation_layout.addStretch()
        
        self.strict_validation_checkbox = QCheckBox()
        self.strict_validation_checkbox.setChecked(self.strict_validation_enabled)
        self.strict_validation_checkbox.setStyleSheet("""
            QCheckBox::indicator {
                width: 20px;
                height: 20px;
            }
            QCheckBox::indicator:unchecked {
                border: 2px solid #94a3b8;
                border-radius: 4px;
                background-color: white;
            }
            QCheckBox::indicator:checked {
                border: 2px solid #3b82f6;
                border-radius: 4px;
                background-color: #3b82f6;
                image: url(none);
            }
            QCheckBox::indicator:checked:hover {
                background-color: #2563eb;
            }
        """)
        validation_layout.addWidget(self.strict_validation_checkbox)
        
        layout.addWidget(validation_widget)
        
        # Описание
        desc_label = QLabel("При включении проверяются только штрихкоды, соответствующие стандартам Wildberries и Ozon")
        desc_label.setStyleSheet("color: #64748b; font-size: 11px; margin-left: 35px; margin-bottom: 10px;")
        desc_label.setWordWrap(True)
        layout.addWidget(desc_label)
        
        layout.addStretch()
        
        # Кнопки
        buttons_widget = QWidget()
        buttons_layout = QHBoxLayout(buttons_widget)
        buttons_layout.setContentsMargins(0, 0, 0, 0)
        buttons_layout.setSpacing(10)
        
        buttons_layout.addStretch()
        
        save_button = QPushButton("✅ Сохранить")
        save_button.setFixedSize(120, 40)
        save_button.setStyleSheet("""
            QPushButton {
                background-color: #3b82f6;
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #2563eb;
            }
        """)
        save_button.clicked.connect(lambda: self.save_settings(dialog))
        buttons_layout.addWidget(save_button)
    
        cancel_button = QPushButton("✕ Отмена")
        cancel_button.setFixedSize(120, 40)
        cancel_button.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        cancel_button.clicked.connect(dialog.reject)
        buttons_layout.addWidget(cancel_button)
    
        layout.addWidget(buttons_widget)
    
        dialog.exec_()

    def save_settings(self, dialog):
        self.strict_validation_enabled = self.strict_validation_checkbox.isChecked()
        self.save_state()
        dialog.accept()
        
        # Показываем уведомление
        self.scan_notification.show_notification("✅ Настройки сохранены")

    def create_search_frame(self):
        self.search_frame = QWidget()
        self.main_layout.addWidget(self.search_frame)
        search_layout = QHBoxLayout()
        search_layout.setContentsMargins(5,5,5,0)
        self.search_frame.setLayout(search_layout)

        self.search_label = QLabel("🔍 Поиск:")
        search_layout.addWidget(self.search_label)

        self.search_entry = QLineEdit()
        search_layout.addWidget(self.search_entry)
        self.search_entry.textChanged.connect(self.filter_items)
        self.search_entry.setContextMenuPolicy(Qt.CustomContextMenu)
        self.search_entry.customContextMenuRequested.connect(lambda event: self.show_paste_menu(event, self.search_entry))
        self.search_entry.setMaximumWidth(300)

        tooltip_search_entry = ToolTip(self.search_entry)
        tooltip_search_entry.setToolTip("Введите текст для фильтрации списка товаров")

        spacer = QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum)
        search_layout.addItem(spacer)

    def create_box_frame(self):
        self.box_frame = QGroupBox("📦 Короб")
        self.main_layout.addWidget(self.box_frame)
        box_layout = QGridLayout()
        self.box_frame.setLayout(box_layout)

        self.box_label = QLabel("Штрихкод короба:")
        box_layout.addWidget(self.box_label, 0, 0, 1, 1, Qt.AlignLeft)

        self.box_entry = QLineEdit()
        box_layout.addWidget(self.box_entry, 0, 1, 1, 1)
        self.box_entry.setMinimumWidth(200)
        self.box_entry.returnPressed.connect(self.process_box_barcode)
        self.box_entry.setContextMenuPolicy(Qt.CustomContextMenu)
        self.box_entry.customContextMenuRequested.connect(lambda event: self.show_paste_menu(event, self.box_entry))
        self.box_entry.setFocus()

        self.new_box_button = QPushButton("+ Новый короб")
        box_layout.addWidget(self.new_box_button, 0, 2, 1, 1)
        self.new_box_button.clicked.connect(self.new_box)

        tooltip_new_box = ToolTip(self.new_box_button)
        tooltip_new_box.setToolTip("Начать работу с новым коробом (Ctrl+N)")

        tooltip_box_entry = ToolTip(self.box_entry)
        tooltip_box_entry.setToolTip("Введите или отсканируйте штрихкод короба")

        box_layout.setColumnStretch(1, 1)

    def create_item_scan_frame(self):
        self.item_scan_frame = QGroupBox("📷 Сканирование товаров")
        self.main_layout.addWidget(self.item_scan_frame)
        item_scan_layout = QGridLayout()
        self.item_scan_frame.setLayout(item_scan_layout)

        self.item_scan_label = QLabel("Штрихкод товара:")
        item_scan_layout.addWidget(self.item_scan_label, 0, 0, 1, 1, Qt.AlignLeft)

        self.item_scan_entry = QLineEdit()
        item_scan_layout.addWidget(self.item_scan_entry, 0, 1, 1, 1)
        self.item_scan_entry.setMinimumWidth(200)
        self.item_scan_entry.returnPressed.connect(self.process_item_barcode)
        self.item_scan_entry.setContextMenuPolicy(Qt.CustomContextMenu)
        self.item_scan_entry.customContextMenuRequested.connect(lambda event: self.show_paste_menu(event, self.item_scan_entry))
        self.item_scan_entry.setEnabled(False)

        tooltip_item_entry = ToolTip(self.item_scan_entry)
        tooltip_item_entry.setToolTip("Введите или отсканируйте штрихкод товара")

        item_scan_layout.setColumnStretch(1, 1)

        self.autoclear_item_entry = QCheckBox("🧹 Очищать поле ввода")
        self.autoclear_item_entry.setChecked(True)
        item_scan_layout.addWidget(self.autoclear_item_entry, 0, 2, 1, 1, Qt.AlignLeft)

        tooltip_autoclear = ToolTip(self.autoclear_item_entry)
        tooltip_autoclear.setToolTip("Автоматически очищать поле ввода после каждого сканирования")

    def create_items_frame(self):
        self.items_frame = QGroupBox("📋 Товары")
        self.main_layout.addWidget(self.items_frame)
        items_layout = QVBoxLayout()
        self.items_frame.setLayout(items_layout)

        self.items_tree = QTreeWidget()
        items_layout.addWidget(self.items_tree)
        self.items_tree.setColumnCount(6)
        self.items_tree.setHeaderLabels(["Статус", "📦 Короб", "🏷 Товар", "🛒 Собрано", "📋 План", "💬 Комментарий"])
        self.items_tree.header().setSectionResizeMode(QHeaderView.Interactive)
        self.items_tree.setAlternatingRowColors(True)
        self.items_tree.itemClicked.connect(self.clear_selection)
        self.items_tree.customContextMenuRequested.connect(self.show_context_menu)
        self.items_tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.items_tree.itemDoubleClicked.connect(self.on_double_click)
        self.items_tree.header().sectionResized.connect(self.save_column_settings)
        for i in range(self.items_tree.columnCount()):
            self.items_tree.headerItem().setTextAlignment(i, Qt.AlignCenter)
        self.items_tree.setColumnWidth(0, 80)
        self.items_tree.setColumnWidth(1, 180)
        self.items_tree.setColumnWidth(2, 180)
        self.items_tree.setColumnWidth(3, 150)
        self.items_tree.setColumnWidth(4, 100)
        self.items_tree.setColumnWidth(5, 200)

    def create_stats_frame(self):
        self.stats_frame = QGroupBox("📊 Статистика сборки")
        self.main_layout.addWidget(self.stats_frame)
        stats_layout = QGridLayout()
        self.stats_frame.setLayout(stats_layout)
        
        time_layout = QHBoxLayout()
        self.time_label = QLabel("⏱️ Время: 00:00:00")
        self.time_label.setStyleSheet("font-size: 11pt; font-weight: bold; color: #2c3e50;")
        time_layout.addWidget(self.time_label)
        
        self.pause_button = QPushButton("⏸")
        self.pause_button.setFixedWidth(50)
        self.pause_button.clicked.connect(self.toggle_pause)
        self.pause_button.setToolTip("Пауза")
        self.pause_button.hide()
        time_layout.addWidget(self.pause_button)
        time_layout.addStretch()
        
        stats_layout.addLayout(time_layout, 0, 0)
        
        self.speed_label = QLabel("⚡ Скорость: 0/мин")
        self.speed_label.setStyleSheet("font-size: 11pt; font-weight: bold; color: #2c3e50;")
        stats_layout.addWidget(self.speed_label, 0, 1)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("%p% (%v/%m)")
        self.progress_bar.setStyleSheet("QProgressBar { height: 20px; }")
        stats_layout.addWidget(self.progress_bar, 1, 0, 1, 2)
        
        status_layout = QHBoxLayout()
        
        self.match_label = QLabel("✅ 0")
        self.match_label.setStyleSheet("color: #27ae60; font-weight: bold;")
        status_layout.addWidget(self.match_label)
        
        self.shortage_label = QLabel("⚠️ 0 (0)")
        self.shortage_label.setStyleSheet("color: #f39c12; font-weight: bold;")
        status_layout.addWidget(self.shortage_label)
        
        self.excess_label = QLabel("❗ 0 (+0)")
        self.excess_label.setStyleSheet("color: #e67e22; font-weight: bold;")
        status_layout.addWidget(self.excess_label)
        
        self.extra_label = QLabel("❓ 0 (+0)")
        self.extra_label.setStyleSheet("color: #7f8c8d; font-weight: bold;")
        status_layout.addWidget(self.extra_label)
        
        status_layout.addStretch()
        stats_layout.addLayout(status_layout, 2, 0, 1, 2)

    def toggle_pause(self):
        if not self.invoice_loaded:
            return
            
        self.is_paused = not self.is_paused
        if self.is_paused:
            self.pause_button.setText("▶️")
            self.pause_button.setToolTip("Старт")
            self.pause_start = time()
        else:
            self.pause_button.setText("⏸️")
            self.pause_button.setToolTip("Пауза")
            if self.pause_start and self.start_time:
                paused_duration = time() - self.pause_start
                self.start_time += paused_duration
            self.pause_start = None
        self.save_state()

    def create_control_frame(self):
        self.control_frame = QWidget()
        self.main_layout.addWidget(self.control_frame)
        control_layout = QHBoxLayout()
        self.control_frame.setLayout(control_layout)

        self.history_button = QPushButton("📜 История")
        control_layout.addWidget(self.history_button)
        self.history_button.clicked.connect(self.show_history)

        tooltip_history_button = ToolTip(self.history_button)
        tooltip_history_button.setToolTip("Открыть окно истории сканирования")

        self.invoice_button = QPushButton("📋 Загрузить накладную")
        control_layout.addWidget(self.invoice_button)
        self.invoice_button.clicked.connect(self.load_invoice_dialog)
        
        invoice_tooltip = "📋 Формат накладной:\n• Файл Excel (.xlsx, .xls)\n• Столбец A: штрихкод\n• Столбец B: количество\n• Первая строка - заголовки (пропускаются)"
        self.invoice_button.setToolTip(invoice_tooltip)
        ToolTip(self.invoice_button).setToolTip(invoice_tooltip)
        
        self.view_invoice_button = QPushButton("👁️ Просмотр накладной")
        control_layout.addWidget(self.view_invoice_button)
        self.view_invoice_button.clicked.connect(self.view_invoice)
        self.view_invoice_button.setEnabled(False)
        self.view_invoice_button.setToolTip("Показать содержимое загруженной накладной")
        
        self.clear_invoice_button = QPushButton("❌ Сбросить накладную")
        control_layout.addWidget(self.clear_invoice_button)
        self.clear_invoice_button.clicked.connect(self.clear_invoice)
        self.clear_invoice_button.setEnabled(False)

        self.undo_button = QPushButton("↩️ Отмена")
        self.undo_button.setObjectName("undo_button")
        self.undo_button.clicked.connect(self.undo_last_action)
        self.undo_button.setEnabled(False)
        self.undo_button.setToolTip("Отменить последнее действие (Ctrl+Z)")
        control_layout.addWidget(self.undo_button)

        self.summary_label = QLabel("")
        control_layout.addWidget(self.summary_label)

        self.save_button = QPushButton("💾 Сохранить")
        control_layout.addWidget(self.save_button)
        self.save_button.clicked.connect(self.save_with_format_dialog)
        self.save_button.setEnabled(False)
        
        control_layout.addStretch()

    def create_status_bar(self):
        self.status_bar = self.statusBar()
        self.status_bar.setStyleSheet(f"QStatusBar{{background-color: {self.COLOR_HEADER_BG}; border-top: 1px solid #ced4da;}}")
        self.status_bar.showMessage("💡 Перетащите CSV или Excel файл в окно для быстрого импорта")

    def get_total_scanned_for_item(self, item_barcode, exclude_box=None):
        total = 0
        for box_barcode, items in self.all_boxes.items():
            if exclude_box and box_barcode == exclude_box:
                continue
            if item_barcode in items:
                total += items[item_barcode]
        return total

    def update_stats(self):
        if not self.invoice_loaded:
            self.time_label.setText("⏱️ Время: 00:00:00")
            self.speed_label.setText("⚡ Скорость: 0/мин")
            self.progress_bar.setValue(0)
            return
            
        if not self.first_scan_done or self.is_paused:
            return
            
        if self.start_time:
            elapsed = int(time() - self.start_time)
            hours = elapsed // 3600
            minutes = (elapsed % 3600) // 60
            seconds = elapsed % 60
            self.time_label.setText(f"⏱️ Время: {hours:02d}:{minutes:02d}:{seconds:02d}")
        
        if self.total_scans > 0 and self.start_time and not self.is_paused:
            elapsed = time() - self.start_time
            if elapsed > 0:
                speed = (self.total_scans * 60) / elapsed
                self.speed_label.setText(f"⚡ Скорость: {speed:.1f}/мин")
        
        if self.invoice_loaded:
            total_planned = sum(self.invoice_data.values())
            total_scanned_planned = 0
            total_scanned_all = 0
            
            for item_barcode, planned in self.invoice_data.items():
                scanned = self.get_total_scanned_for_item(item_barcode)
                total_scanned_planned += min(scanned, planned)
                total_scanned_all += scanned
            
            for box_barcode, items in self.all_boxes.items():
                for item_barcode, count in items.items():
                    if item_barcode not in self.invoice_data:
                        total_scanned_all += count
            
            match_count = 0
            shortage_count = 0
            excess_count = 0
            extra_count = 0
            
            shortage_units = 0
            excess_units = 0
            extra_units = 0
            
            for item_barcode, planned in self.invoice_data.items():
                scanned = self.get_total_scanned_for_item(item_barcode)
                
                if scanned == planned:
                    match_count += 1
                elif scanned < planned:
                    shortage_count += 1
                    shortage_units += planned - scanned
                else:
                    excess_count += 1
                    excess_units += scanned - planned
            
            for box_barcode, items in self.all_boxes.items():
                for item_barcode, count in items.items():
                    if item_barcode not in self.invoice_data:
                        extra_count += 1
                        extra_units += count
            
            if total_planned > 0:
                progress = min(100, int((total_scanned_planned / total_planned) * 100))
                self.progress_bar.setValue(progress)
                self.progress_bar.setFormat(f"%p% ({total_scanned_planned}/{total_planned})")
                
                if total_scanned_all > total_planned:
                    self.progress_bar.setStyleSheet("QProgressBar::chunk { background-color: #e74c3c; }")
                else:
                    self.progress_bar.setStyleSheet("QProgressBar::chunk { background-color: #3498db; }")
            
            self.match_label.setText(f"✅ {match_count}")
            self.shortage_label.setText(f"⚠️ {shortage_count} (-{shortage_units})")
            self.excess_label.setText(f"❗ {excess_count} (+{excess_units})")
            self.extra_label.setText(f"❓ {extra_count} (+{extra_units})")

    def convert_ru_to_en_layout_box(self, barcode):
        if len(barcode) >= 3 and barcode.lower().startswith('ца'):
            barcode = 'wb' + barcode[2:]
        elif len(barcode) >= 4 and barcode.lower().startswith('ци_'):
            barcode = 'WB_' + barcode[3:]
        return barcode

    def convert_ru_to_en_layout_item(self, barcode):
        if len(barcode) >= 4 and barcode.lower().startswith('щят'):
            barcode = 'OZN' + barcode[3:]
        return barcode

    def is_valid_barcode(self, barcode, barcode_type):
        if not self.strict_validation_enabled:
            return bool(re.match(r"^[\w\-\./]+$", barcode)) and 4 <= len(barcode) <= 50
        
        patterns = {
            'box': [
                r'^WB_[\w\-]+$',
                r'^\d{8,}$',
                r'^[A-Z]{2}\d{6,}$',
                r'^[A-Z0-9]{10,}$'
            ],
            'item': [
                r'^\d{8}$',
                r'^\d{12}$',
                r'^\d{13}$',
                r'^OZN\d+$',
                r'^ozn\d+$',
                r'^[A-Z]{2}\d{9}[A-Z]{2}$',
                r'^[0-9]{8,14}$'
            ]
        }
        
        for pattern in patterns.get(barcode_type, patterns['item']):
            if re.match(pattern, barcode, re.IGNORECASE):
                return 4 <= len(barcode) <= 50
        return False

    def check_duplicate_item(self, barcode, current_box):
        """Проверка дубликатов с учетом плана"""
        if not self.invoice_loaded:
            return []  # Без накладной не предупреждаем
            
        total_scanned = self.get_total_scanned_for_item(barcode)
        planned = self.invoice_data.get(barcode, 0)
        
        # Если план есть и общее количество не превышает план - ок
        if planned > 0 and total_scanned < planned:
            return []
            
        # Ищем в каких коробах уже есть этот товар
        duplicate_boxes = []
        for box_barcode, items in self.all_boxes.items():
            if box_barcode != current_box and barcode in items:
                duplicate_boxes.append(box_barcode)
        return duplicate_boxes

    def update_undo_button_state(self):
        self.undo_button.setEnabled(self.undo_manager.can_undo())

    def undo_last_action(self):
        if not self.undo_manager.can_undo():
            return
            
        action = self.undo_manager.undo()
        if not action:
            return
            
        action_type = action.get('type')
        
        if action_type == 'scan':
            box = action['box_barcode']
            item = action['barcode']
            
            if box in self.all_boxes and item in self.all_boxes[box]:
                old_count = self.all_boxes[box][item]
                
                if self.all_boxes[box][item] <= 1:
                    del self.all_boxes[box][item]
                    if not self.all_boxes[box]:
                        del self.all_boxes[box]
                        if self.current_box_barcode == box:
                            self.current_box_barcode = ""
                            self.box_entry.setEnabled(True)
                            self.item_scan_entry.setEnabled(False)
                else:
                    self.all_boxes[box][item] -= 1
                
                self.scan_history.append({
                    'timestamp': datetime.now().isoformat(),
                    'type': 'item',
                    'barcode': item,
                    'box_barcode': box,
                    'action': 'undo',
                    'action_type': 'undo',
                    'details': f'↩️ Отмена сканирования (было {old_count} → {old_count - 1})'
                })
                
                self.total_scans = max(0, self.total_scans - 1)
                self.update_status(f"↩️ Отменён товар: {item}")
                self.scan_notification.show_notification(f"↩️ Отмена: {item}")
                
        elif action_type == 'edit_count':
            box = action['box_barcode']
            item = action['barcode']
            old_count = action['old_value']
            new_count = action['new_value']
            
            if box in self.all_boxes:
                if old_count == 0:
                    # Было удаление - восстанавливаем
                    self.all_boxes[box][item] = new_count
                else:
                    self.all_boxes[box][item] = old_count
                    
            self.scan_history.append({
                'timestamp': datetime.now().isoformat(),
                'type': 'item',
                'barcode': item,
                'box_barcode': box,
                'action': 'undo',
                'action_type': 'undo',
                'details': f'↩️ Отмена изменения количества: {new_count} → {old_count}'
            })
            
        elif action_type == 'edit_barcode':
            box = action['box_barcode']
            old_barcode = action['old_value']
            new_barcode = action['new_value']
            count = action['count']
            
            if box in self.all_boxes and new_barcode in self.all_boxes[box]:
                del self.all_boxes[box][new_barcode]
                self.all_boxes[box][old_barcode] = count
                
            self.scan_history.append({
                'timestamp': datetime.now().isoformat(),
                'type': 'item',
                'barcode': old_barcode,
                'box_barcode': box,
                'action': 'undo',
                'action_type': 'undo',
                'details': f'↩️ Отмена изменения штрихкода: {new_barcode} → {old_barcode}'
            })
            
        self.has_unsaved_changes = True
        self.refresh_treeview()
        self.update_undo_button_state()

    def process_box_barcode(self):
        barcode_input = self.box_entry.text().strip()
        barcode = self.convert_ru_to_en_layout_box(barcode_input)
        
        if not barcode:
            self.show_warning("Введите штрихкод короба!")
            return

        if not self.is_valid_barcode(barcode, barcode_type='box'):
            self.show_error("Неверный штрихкод короба!")
            self.box_entry.clear()
            return

        if barcode not in self.all_boxes:
            self.all_boxes[barcode] = {}

        self.current_box_barcode = barcode
        self.box_entry.setEnabled(False)
        self.item_scan_entry.setEnabled(True)
        self.item_scan_entry.setFocus()
        self.save_button.setEnabled(True)
        
        if self.invoice_loaded and not self.first_scan_done and not self.start_time:
            self.start_time = time()
            self.first_scan_done = True
            self.pause_button.show()
            
        self.has_unsaved_changes = True
        self.update_status(f"✅ Текущий короб: {self.current_box_barcode}")
        self.refresh_treeview()
        
        self.scan_history.append({
            'timestamp': datetime.now().isoformat(),
            'type': 'box',
            'barcode': barcode,
            'action': 'scan',
            'action_type': 'scan',
            'details': ''
        })
        
        self.highlight_entry(self.box_entry)
        self.scan_notification.show_notification(f"📦 Короб: {barcode}")
        self.update_undo_button_state()

    def process_item_barcode(self):
        barcode_input = self.item_scan_entry.text().strip()
        barcode = self.convert_ru_to_en_layout_item(barcode_input)

        if not self.current_box_barcode:
            self.show_warning("Сначала отсканируйте штрихкод короба!")
            self.item_scan_entry.clear()
            self.box_entry.setFocus()
            return
        if not barcode:
            self.show_warning("Введите штрихкод товара!")
            return
        if not self.is_valid_barcode(barcode, barcode_type='item'):
            self.show_error("Неверный штрихкод товара!")
            self.item_scan_entry.clear()
            return
        if self.current_box_barcode not in self.all_boxes:
            QMessageBox.critical(self, "Ошибка", "Текущий короб не найден!")
            return
        
        duplicate_boxes = self.check_duplicate_item(barcode, self.current_box_barcode)
        if duplicate_boxes:
            total_scanned = self.get_total_scanned_for_item(barcode)
            planned = self.invoice_data.get(barcode, 0)
            
            msg = f"⚠️ Товар {barcode} уже есть в короб(ах):\n"
            for box in duplicate_boxes:
                msg += f"  • {box}\n"
            
            if planned > 0:
                msg += f"\n📊 План: {planned}, всего собрано (с учётом этого): {total_scanned + 1}"
                if total_scanned + 1 > planned:
                    msg += f"\n❗ Будет ПЕРЕБОР на {total_scanned + 1 - planned}"
                else:
                    msg += f"\n✅ В пределах плана"
            
            msg += "\n\nПродолжить сканирование?"
            
            dialog = ConfirmationDialog(
                "⚠️ Проверка дубликатов",
                msg,
                "question",
                self
            )
            dialog.yes_button.setText("✅ Да")
            dialog.no_button.setText("✕ Нет")
            
            if dialog.exec_() == QDialog.Rejected:
                self.item_scan_entry.clear()
                return
            
        self.add_item(barcode)
        
        self.total_scans += 1
        self.has_unsaved_changes = True
        
        if self.invoice_loaded and not self.first_scan_done and not self.start_time:
            self.start_time = time()
            self.first_scan_done = True
            self.pause_button.show()
        
        # Добавляем действие в стек отмены
        self.undo_manager.add_action({
            'type': 'scan',
            'barcode': barcode,
            'box_barcode': self.current_box_barcode
        })
        
        self.scan_history.append({
            'timestamp': datetime.now().isoformat(),
            'type': 'item',
            'barcode': barcode,
            'box_barcode': self.current_box_barcode,
            'action': 'scan',
            'action_type': 'scan',
            'details': f'📷 Сканирование товара'
        })
        
        if self.invoice_loaded:
            if barcode not in self.invoice_data:
                self.scan_notification.show_notification(f"⚠️ Товар {barcode}\nотсутствует в накладной!", True)
                QApplication.beep()
            else:
                total_scanned = self.get_total_scanned_for_item(barcode)
                planned = self.invoice_data[barcode]
                
                if total_scanned > planned:
                    self.scan_notification.show_notification(f"❗ ПЕРЕБОР: {barcode}\nплан: {planned}, всего: {total_scanned}", True)
                    QApplication.beep()
                elif total_scanned == planned:
                    self.scan_notification.show_notification(f"✅ План выполнен: {barcode}")
                else:
                    remaining = planned - total_scanned
                    self.scan_notification.show_notification(f"✅ {barcode}\nосталось: {remaining}")
        
        if self.autoclear_item_entry.isChecked():
            self.item_scan_entry.clear()
        self.highlight_entry(self.item_scan_entry)
        self.save_state()
        self.update_undo_button_state()

    def show_history(self):
        if self.history_window and self.history_window.isVisible():
            self.history_window.raise_()
            self.history_window.activateWindow()
            return

        self.history_window = QDialog(self)
        self.history_window.setWindowTitle("📜 История сканирования")
        self.history_window.setGeometry(100, 100, 1200, 700)
        layout = QVBoxLayout(self.history_window)

        filter_frame = QWidget()
        layout.addWidget(filter_frame)
        filter_layout = QHBoxLayout(filter_frame)

        filter_label = QLabel("🔍 Фильтр:")
        filter_layout.addWidget(filter_label)

        self.history_filter_entry = QLineEdit()
        filter_layout.addWidget(self.history_filter_entry)
        self.history_filter_entry.textChanged.connect(self.filter_history)
        
        filter_layout.addStretch()
        
        info_label = QLabel(f"👤 Сборщик: {self.packer_name if self.packer_name else 'не указан'}")
        info_label.setStyleSheet("color: #3498db; font-weight: bold;")
        filter_layout.addWidget(info_label)

        self.history_tree = QTreeWidget()
        layout.addWidget(self.history_tree)
        self.history_tree.setColumnCount(6)
        self.history_tree.setHeaderLabels(["📦 Короб", "🏷 Товар", "📦 Кол-во", "⚡ Действие", "📝 Детали", "📅 Время"])
        self.history_tree.header().setSectionResizeMode(QHeaderView.Stretch)
        self.history_tree.header().setSectionResizeMode(0, QHeaderView.Interactive)
        self.history_tree.header().setSectionResizeMode(1, QHeaderView.Interactive)
        self.history_tree.header().setSectionResizeMode(2, QHeaderView.Interactive)
        self.history_tree.header().setSectionResizeMode(3, QHeaderView.Interactive)
        self.history_tree.header().setSectionResizeMode(4, QHeaderView.Interactive)
        self.history_tree.header().setSectionResizeMode(5, QHeaderView.Interactive)
        self.history_tree.setColumnWidth(0, 200)
        self.history_tree.setColumnWidth(1, 180)
        self.history_tree.setColumnWidth(2, 120)
        self.history_tree.setColumnWidth(3, 140)
        self.history_tree.setColumnWidth(4, 300)
        self.history_tree.setAlternatingRowColors(True)

        self.populate_history_tree()
        self.history_window.show()

    def populate_history_tree(self):
        self.history_tree.clear()
        
        boxes_data = {}
        
        for entry in self.scan_history:
            if entry['type'] == 'box':
                box_barcode = entry['barcode']
                if box_barcode not in boxes_data:
                    boxes_data[box_barcode] = {
                        'timestamp': entry['timestamp'],
                        'action': entry.get('action', 'scan'),
                        'action_type': entry.get('action_type', 'scan'),
                        'details': entry.get('details', ''),
                        'items': {}
                    }
        
        for entry in self.scan_history:
            if entry['type'] == 'item':
                box_barcode = entry.get('box_barcode')
                if box_barcode in boxes_data:
                    item_barcode = entry['barcode']
                    if item_barcode not in boxes_data[box_barcode]['items']:
                        boxes_data[box_barcode]['items'][item_barcode] = []
                    boxes_data[box_barcode]['items'][item_barcode].append(entry)

        for box_barcode, box_data in boxes_data.items():
            try:
                dt = datetime.fromisoformat(box_data['timestamp'])
                box_time = dt.strftime("%d.%m.%Y %H:%M:%S")
            except:
                box_time = box_data['timestamp']
            
            action_type_display = "📌" if box_data['action_type'] == 'scan' else "✏️"
            action_display = "Открытие короба" if box_data['action_type'] == 'scan' else "Изменение короба"
            
            box_item = QTreeWidgetItem(self.history_tree)
            box_item.setText(0, f"📦 {box_barcode}")
            box_item.setText(1, "")
            box_item.setText(2, action_type_display)
            box_item.setText(3, action_display)
            box_item.setText(4, box_data['details'])
            box_item.setText(5, box_time)
            
            font = QFont()
            font.setBold(True)
            for i in range(6):
                box_item.setFont(i, font)
            
            for item_barcode, entries in box_data['items'].items():
                # Получаем актуальное количество товара в коробе
                actual_count = self.all_boxes.get(box_barcode, {}).get(item_barcode, 0)
                
                item_main = QTreeWidgetItem(box_item)
                item_main.setText(0, "  🏷")
                item_main.setText(1, item_barcode)
                item_main.setText(2, f"📦 {actual_count}")
                item_main.setText(3, f"{len(entries)} операций")
                item_main.setText(4, "")
                item_main.setText(5, "")
                
                font = QFont()
                font.setBold(True)
                for i in range(6):
                    item_main.setFont(i, font)
                
                for entry in entries:
                    try:
                        dt = datetime.fromisoformat(entry['timestamp'])
                        item_time = dt.strftime("%d.%m.%Y %H:%M:%S")
                    except:
                        item_time = entry['timestamp']
                    
                    action_type = entry.get('action_type', 'scan')
                    action = entry.get('action', 'scan')
                    details = entry.get('details', '')
                    
                    if action_type == 'undo':
                        action_type_icon = "↩️"
                    elif action_type == 'edit':
                        action_type_icon = "✏️"
                    else:
                        action_type_icon = "📷"
                    
                    if action == 'scan':
                        action_text = "Сканирование"
                    elif action == 'edit_count':
                        action_text = "Изменение количества"
                    elif action == 'edit_barcode':
                        action_text = "Изменение штрихкода"
                    elif action == 'edit_comment':
                        action_text = "Изменение комментария"
                    elif action == 'delete':
                        action_text = "Удаление"
                    elif action == 'undo':
                        action_text = "↩️ Отмена"
                    else:
                        action_text = action
                    
                    entry_item = QTreeWidgetItem(item_main)
                    entry_item.setText(0, "    •")
                    entry_item.setText(1, "")
                    entry_item.setText(2, action_type_icon)
                    entry_item.setText(3, action_text)
                    entry_item.setText(4, details)
                    entry_item.setText(5, item_time)
                    
                    if action_type == 'edit':
                        for i in range(6):
                            entry_item.setForeground(i, QColor("#e67e22"))
                    elif action_type == 'undo':
                        for i in range(6):
                            entry_item.setForeground(i, QColor("#9b59b6"))
                    else:
                        for i in range(6):
                            entry_item.setForeground(i, QColor("#2ecc71"))
                
                item_main.setExpanded(False)
            
            box_item.setExpanded(False)

    def filter_history(self):
        if not self.history_tree:
            return
            
        filter_text = self.history_filter_entry.text().lower()
        
        for i in range(self.history_tree.topLevelItemCount()):
            box_item = self.history_tree.topLevelItem(i)
            box_visible = False
            
            if filter_text in box_item.text(0).lower():
                box_visible = True
            
            for j in range(box_item.childCount()):
                item_main = box_item.child(j)
                item_visible = False
                
                if filter_text in item_main.text(1).lower():
                    item_visible = True
                    box_visible = True
                
                for k in range(item_main.childCount()):
                    entry_item = item_main.child(k)
                    entry_visible = False
                    
                    if (filter_text in entry_item.text(3).lower() or 
                        filter_text in entry_item.text(4).lower()):
                        entry_visible = True
                        item_visible = True
                        box_visible = True
                    
                    entry_item.setHidden(not entry_visible)
                
                item_main.setHidden(not item_visible)
            
            box_item.setHidden(not box_visible)

    def highlight_entry(self, entry):
        entry.setStyleSheet("QLineEdit { background-color: #c8e6c9; }")
        QTimer.singleShot(200, lambda: entry.setStyleSheet(""))

    def add_item(self, item_barcode):
        if item_barcode in self.all_boxes[self.current_box_barcode]:
            self.all_boxes[self.current_box_barcode][item_barcode] += 1
        else:
            self.all_boxes[self.current_box_barcode][item_barcode] = 1
        self.refresh_treeview()

    def refresh_treeview(self):
        self.items_tree.clear()
        for box_barcode, items in self.all_boxes.items():
            box_comment = self.comments.get((box_barcode, ""), "")
            box_item = QTreeWidgetItem(self.items_tree, ["", box_barcode, "", "", "", box_comment])
            box_item.setFlags(box_item.flags() | Qt.ItemIsTristate)
            
            font = QFont()
            font.setBold(True)
            for i in range(6):
                box_item.setFont(i, font)
            
            self.items_tree.expandItem(box_item)

            for item_barcode, count in items.items():
                item_comment = self.comments.get((box_barcode, item_barcode), "")
                if not self.search_query or self.search_query.lower() in box_barcode.lower() or self.search_query.lower() in item_barcode.lower():
                    status_icon = ""
                    planned = ""
                    
                    if self.invoice_loaded:
                        if item_barcode in self.invoice_data:
                            planned = str(self.invoice_data[item_barcode])
                            total_scanned = self.get_total_scanned_for_item(item_barcode)
                            planned_int = self.invoice_data[item_barcode]
                            
                            if total_scanned == planned_int:
                                status_icon = "✅"
                            elif total_scanned < planned_int:
                                status_icon = "⚠️"
                            else:
                                status_icon = "❗"
                        else:
                            status_icon = "❓"
                            planned = "0"
                    
                    item = QTreeWidgetItem(box_item, [status_icon, "", item_barcode, str(count), planned, item_comment])
                    for i in range(1, 6):
                        item.setTextAlignment(i, Qt.AlignCenter)
                    
                    if status_icon == "❓":
                        for i in range(6):
                            item.setForeground(i, QBrush(QColor("#e67e22")))
                            item.setBackground(i, QBrush(QColor("#fff3e0")))
        self.update_summary()
        self.update_stats()

    def filter_items(self):
        self.search_query = self.search_entry.text()
        self.refresh_treeview()

    def show_context_menu(self, point):
        try:
            item = self.items_tree.itemAt(point)
            if item is None:
                return

            self.items_tree.setCurrentItem(item)
            column_index = -1

            header = self.items_tree.header()
            x_click = point.x()

            for i in range(self.items_tree.columnCount()):
                section_pos = header.sectionViewportPosition(i)
                section_width = header.sectionSize(i)
                if x_click >= section_pos and x_click < section_pos + section_width:
                    column_index = i
                    break

            values = [item.text(i) for i in range(self.items_tree.columnCount())]
        
            # Проверяем что values достаточно
            if len(values) < 6:
                return

            context_menu = QMenu(self)
            context_menu.setStyleSheet("""
                QMenu {
                    background-color: white;
                    border: 1px solid #e2e8f0;
                    border-radius: 12px;
                    padding: 8px 4px;
                    font-size: 13px;
                    color: #1e293b;
                    box-shadow: 0 4px 12px rgba(0, 0, 0, 0.15);
                }
                QMenu::item {
                    padding: 8px 24px 8px 12px;
                    border-radius: 6px;
                    margin: 2px 4px;
                }
                QMenu::item:selected {
                    background-color: #f1f5f9;
                    color: #3b82f6;
                }
                QMenu::item:selected:!enabled {
                    background-color: transparent;
                    color: #94a3b8;
                }
                QMenu::icon {
                    padding-left: 8px;
                }
                QMenu::separator {
                    height: 1px;
                    background-color: #e2e8f0;
                    margin: 8px 12px;
                }
            """)

            # Проверяем короб ли это (все поля кроме статуса и комментария пустые)
            is_box = (len(values) == 6 and 
                     values[2] == "" and 
                     values[3] == "" and 
                     values[4] == "")

            if is_box:
                # Для короба
                if column_index == 1:
                    action_copy_box_barcode = QAction("📋 Копировать штрихкод короба", self)
                    action_copy_box_barcode.triggered.connect(lambda: self.clipboard.setText(values[1]))
                    context_menu.addAction(action_copy_box_barcode)
                elif column_index == 5:
                    action_edit_comment = QAction("✏️ Изменить комментарий к коробу", self)
                    action_edit_comment.triggered.connect(lambda: self.edit_comment(item))
                    context_menu.addAction(action_edit_comment)

                # Эти пункты показываем всегда для короба
                context_menu.addSeparator()

                action_edit_box_barcode = QAction("✏️ Изменить штрихкод короба", self)
                action_edit_box_barcode.triggered.connect(lambda: self.edit_box_barcode(item))
                context_menu.addAction(action_edit_box_barcode)

                context_menu.addSeparator()

                action_delete_box = QAction("🗑️ Удалить короб", self)
                action_delete_box.triggered.connect(lambda: self.delete_box(item))
                context_menu.addAction(action_delete_box)

            else:
                # Для товара
                parent_item = item.parent()
                if parent_item is None:
                    return
                
                box_barcode = parent_item.text(1) if parent_item else ""
            
                # Проверяем что значения существуют
                item_barcode = values[2] if len(values) > 2 else ""
                count = values[3] if len(values) > 3 else ""
                planned = values[4] if len(values) > 4 else ""

                # Группа копирования
                copy_menu = context_menu.addMenu("📋 Копировать")
                copy_menu.setStyleSheet("""
                    QMenu {
                        background-color: white;
                        border: 1px solid #e2e8f0;
                        border-radius: 8px;
                        padding: 4px;
                    }
                """)
            
                # Копировать штрихкод короба
                action_copy_box = QAction("📦 Штрихкод короба", self)
                action_copy_box.triggered.connect(lambda checked, b=box_barcode: self.clipboard.setText(b))
                copy_menu.addAction(action_copy_box)
            
                # Копировать штрихкод товара
                if item_barcode:
                    action_copy_item = QAction("🏷 Штрихкод товара", self)
                    action_copy_item.triggered.connect(lambda checked, b=item_barcode: self.clipboard.setText(b))
                    copy_menu.addAction(action_copy_item)
            
                # Копировать количество
                if count:
                    action_copy_count = QAction("🔢 Количество", self)
                    action_copy_count.triggered.connect(lambda checked, c=count: self.clipboard.setText(c))
                    copy_menu.addAction(action_copy_count)
            
                # Копировать план (если есть накладная)
                if self.invoice_loaded and planned:
                    action_copy_planned = QAction("📋 План", self)
                    action_copy_planned.triggered.connect(lambda checked, p=planned: self.clipboard.setText(p))
                    copy_menu.addAction(action_copy_planned)

                context_menu.addSeparator()

                # Группа редактирования
                if column_index == 5:
                    action_edit_comment = QAction("✏️ Изменить комментарий", self)
                    action_edit_comment.triggered.connect(lambda: self.edit_comment(item))
                    context_menu.addAction(action_edit_comment)

                if column_index in (2, 3):
                    action_edit_count = QAction("✏️ Изменить количество", self)
                    action_edit_count.triggered.connect(lambda: self.edit_item_count(item))
                    context_menu.addAction(action_edit_count)
            
                if column_index == 2:
                    action_edit_item_barcode = QAction("✏️ Изменить штрихкод товара", self)
                    action_edit_item_barcode.triggered.connect(lambda: self.edit_item_barcode(item))
                    context_menu.addAction(action_edit_item_barcode)

                context_menu.addSeparator()

                # Удаление
                action_delete_item = QAction("🗑️ Удалить товар", self)
                action_delete_item.triggered.connect(lambda: self.delete_item(item))
                context_menu.addAction(action_delete_item)

            context_menu.popup(self.items_tree.viewport().mapToGlobal(point))
        
        except Exception as e:
            print(f"Ошибка в контекстном меню: {e}")
            import traceback
            traceback.print_exc()

    def collapse_box_items(self, box_item):
        for i in range(box_item.childCount()):
            box_item.child(i).setExpanded(False)

    def expand_box_items(self, box_item):
        for i in range(box_item.childCount()):
            box_item.child(i).setExpanded(True)

    def clear_selection(self, item, column):
        if not item.isSelected():
            self.items_tree.clearSelection()

    def edit_item_count(self, selected_item):
        parent_item = selected_item.parent()
        box_barcode = parent_item.text(1) if parent_item else ""
        current_count = selected_item.text(3)
        barcode = selected_item.text(2)

        # Получаем план если есть накладная
        planned = None
        if self.invoice_loaded and barcode in self.invoice_data:
            planned = self.invoice_data[barcode]

        dialog = EditCountDialog(barcode, int(current_count), planned, self)
        if dialog.exec_() == QDialog.Accepted:
            new_count = dialog.get_value()
            old_count = int(current_count)
            
            if str(box_barcode) in self.all_boxes:
                if new_count == 0:
                    if barcode in self.all_boxes[box_barcode]:
                        # Добавляем в стек отмены
                        self.undo_manager.add_action({
                            'type': 'edit_count',
                            'barcode': barcode,
                            'box_barcode': box_barcode,
                            'old_value': old_count,
                            'new_value': 0
                        })
                        
                        del self.all_boxes[box_barcode][barcode]
                        if not self.all_boxes[box_barcode]:
                            del self.all_boxes[box_barcode]
                        
                        self.scan_history.append({
                            'timestamp': datetime.now().isoformat(),
                            'type': 'item',
                            'barcode': barcode,
                            'box_barcode': box_barcode,
                            'action': 'delete',
                            'action_type': 'edit',
                            'details': f'Удаление товара (было {old_count})'
                        })
                else:
                    # Добавляем в стек отмены
                    self.undo_manager.add_action({
                        'type': 'edit_count',
                        'barcode': barcode,
                        'box_barcode': box_barcode,
                        'old_value': old_count,
                        'new_value': new_count
                    })
                    
                    self.all_boxes[str(box_barcode)][barcode] = new_count
                    
                    change = new_count - old_count
                    change_sign = "+" if change > 0 else ""
                    
                    self.scan_history.append({
                        'timestamp': datetime.now().isoformat(),
                        'type': 'item',
                        'barcode': barcode,
                        'box_barcode': box_barcode,
                        'action': 'edit_count',
                        'action_type': 'edit',
                        'details': f'{old_count} → {new_count} ({change_sign}{change})'
                    })
            
            self.has_unsaved_changes = True
            self.refresh_treeview()
            self.update_summary()
            self.update_undo_button_state()
            self.save_state()

    def edit_box_barcode(self, item):
        old_barcode = item.text(1)
        
        # Создаем кастомный диалог
        dialog = QDialog(self)
        dialog.setWindowTitle("✏️ Изменить штрихкод короба")
        dialog.setModal(True)
        dialog.setFixedSize(400, 220)
        dialog.setWindowFlags(Qt.Dialog | Qt.WindowCloseButtonHint)
        
        layout = QVBoxLayout(dialog)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Информация
        info_widget = QWidget()
        info_widget.setStyleSheet("""
            QWidget {
                background-color: #f8fafc;
                border-radius: 8px;
            }
        """)
        info_layout = QHBoxLayout(info_widget)
        info_layout.setContentsMargins(12, 8, 12, 8)
        
        info_icon = QLabel("📦")
        info_icon.setStyleSheet("font-size: 20px; background: transparent;")
        info_layout.addWidget(info_icon)
        
        info_text = QLabel(f"Текущий штрихкод: {old_barcode}")
        info_text.setStyleSheet("font-size: 12px; font-weight: bold; color: #1e293b; background: transparent;")
        info_layout.addWidget(info_text)
        
        layout.addWidget(info_widget)
        
        # Поле ввода
        input_label = QLabel("Новый штрихкод:")
        input_label.setStyleSheet("font-size: 12px; color: #475569; font-weight: bold;")
        layout.addWidget(input_label)
        
        barcode_edit = QLineEdit(old_barcode)
        barcode_edit.setMinimumHeight(40)
        barcode_edit.setStyleSheet("""
            QLineEdit {
                border: 2px solid #cbd5e1;
                border-radius: 8px;
                padding: 8px 12px;
                font-size: 14px;
                font-family: 'Courier New';
            }
            QLineEdit:focus {
                border-color: #3b82f6;
            }
        """)
        layout.addWidget(barcode_edit)
        
        layout.addStretch()
        
        # Кнопки
        buttons_widget = QWidget()
        buttons_layout = QHBoxLayout(buttons_widget)
        buttons_layout.setContentsMargins(0, 0, 0, 0)
        buttons_layout.setSpacing(10)
        
        buttons_layout.addStretch()
        
        ok_button = QPushButton("✅ Применить")
        ok_button.setFixedSize(120, 40)
        ok_button.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)
        ok_button.clicked.connect(dialog.accept)
        buttons_layout.addWidget(ok_button)
        
        cancel_button = QPushButton("✕ Отмена")
        cancel_button.setFixedSize(120, 40)
        cancel_button.setStyleSheet("""
            QPushButton {
               background-color: #e74c3c;
               color: white;
               border: none;
               border-radius: 8px;
               font-weight: bold;
               font-size: 12px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        cancel_button.clicked.connect(dialog.reject)
        buttons_layout.addWidget(cancel_button)
        
        layout.addWidget(buttons_widget)
        
        barcode_edit.setFocus()
        barcode_edit.selectAll()
        barcode_edit.returnPressed.connect(dialog.accept)
        
        if dialog.exec_() == QDialog.Accepted:
            new_barcode = barcode_edit.text().strip()
            if new_barcode and new_barcode != old_barcode:
                if self.is_valid_barcode(new_barcode, barcode_type='box'):
                    if new_barcode not in self.all_boxes:
                        self.all_boxes[new_barcode] = self.all_boxes.pop(old_barcode)
                        for key in list(self.comments.keys()):
                            if key[0] == old_barcode:
                                new_key = (new_barcode, key[1])
                                self.comments[new_key] = self.comments.pop(key)
    
                        if self.current_box_barcode == old_barcode:
                            self.current_box_barcode = new_barcode
                            self.update_status(f"✅ Текущий короб: {self.current_box_barcode}")
                        
                        self.scan_history.append({
                            'timestamp': datetime.now().isoformat(),
                            'type': 'box',
                            'barcode': new_barcode,
                            'action': 'edit_barcode',
                            'action_type': 'edit',
                            'details': f'{old_barcode} → {new_barcode}'
                        })
                        
                        self.has_unsaved_changes = True
                        self.refresh_treeview()
                        self.scan_notification.show_notification(f"✅ Штрихкод изменён")
                    else:
                        self.show_error("Короб с таким штрихкодом уже существует!")
                else:
                    self.show_error("Неверный штрихкод короба!")

    def edit_item_barcode(self, item):
        parent_item = item.parent()
        box_barcode = parent_item.text(1) if parent_item else ""
        old_barcode = item.text(2)
        count = int(item.text(3))
        
        # Создаем кастомный диалог
        dialog = QDialog(self)
        dialog.setWindowTitle("✏️ Изменить штрихкод товара")
        dialog.setModal(True)
        dialog.setFixedSize(400, 300)
        dialog.setWindowFlags(Qt.Dialog | Qt.WindowCloseButtonHint)
        
        layout = QVBoxLayout(dialog)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Информация о коробе
        box_info = QWidget()
        box_info.setStyleSheet("""
            QWidget {
                background-color: #f1f5f9;
                border-radius: 8px;
            }
        """)
        box_layout = QHBoxLayout(box_info)
        box_layout.setContentsMargins(12, 8, 12, 8)
        
        box_icon = QLabel("📦")
        box_icon.setStyleSheet("font-size: 18px; background: transparent;")
        box_layout.addWidget(box_icon)
        
        box_text = QLabel(f"Короб: {box_barcode}")
        box_text.setStyleSheet("font-size: 12px; color: #334155; background: transparent;")
        box_layout.addWidget(box_text)
        
        layout.addWidget(box_info)
        
        # Информация о товаре
        item_info = QWidget()
        item_info.setStyleSheet("""
            QWidget {
                background-color: #f8fafc;
                border-radius: 8px;
            }
        """)
        item_layout = QHBoxLayout(item_info)
        item_layout.setContentsMargins(12, 8, 12, 8)
        
        item_icon = QLabel("🏷")
        item_icon.setStyleSheet("font-size: 18px; background: transparent;")
        item_layout.addWidget(item_icon)
        
        item_text = QLabel(f"Текущий штрихкод: {old_barcode}")
        item_text.setStyleSheet("font-size: 12px; font-weight: bold; color: #1e293b; background: transparent;")
        item_layout.addWidget(item_text)
        
        layout.addWidget(item_info)
        
        # Поле ввода
        input_label = QLabel("Новый штрихкод товара:")
        input_label.setStyleSheet("font-size: 12px; color: #475569; font-weight: bold; margin-top: 5px;")
        layout.addWidget(input_label)
        
        barcode_edit = QLineEdit(old_barcode)
        barcode_edit.setMinimumHeight(40)
        barcode_edit.setStyleSheet("""
            QLineEdit {
                border: 2px solid #cbd5e1;
                border-radius: 8px;
                padding: 8px 12px;
                font-size: 14px;
                font-family: 'Courier New';
            }
            QLineEdit:focus {
                border-color: #3b82f6;
            }
        """)
        layout.addWidget(barcode_edit)
        
        layout.addStretch()
        
        # Кнопки
        buttons_widget = QWidget()
        buttons_layout = QHBoxLayout(buttons_widget)
        buttons_layout.setContentsMargins(0, 0, 0, 0)
        buttons_layout.setSpacing(10)
        
        buttons_layout.addStretch()
        
        ok_button = QPushButton("✅ Применить")
        ok_button.setFixedSize(120, 40)
        ok_button.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                border: none;
                border-radius: 8px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)
        ok_button.clicked.connect(dialog.accept)
        buttons_layout.addWidget(ok_button)
        
        cancel_button = QPushButton("✕ Отмена")
        cancel_button.setFixedSize(120, 40)
        cancel_button.setStyleSheet("""
            QPushButton {
               background-color: #e74c3c;
               color: white;
               border: none;
               border-radius: 8px;
               font-weight: bold;
               font-size: 12px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        cancel_button.clicked.connect(dialog.reject)
        buttons_layout.addWidget(cancel_button)
        
        layout.addWidget(buttons_widget)
        
        barcode_edit.setFocus()
        barcode_edit.selectAll()
        barcode_edit.returnPressed.connect(dialog.accept)
        
        if dialog.exec_() == QDialog.Accepted:
            new_barcode = barcode_edit.text().strip()
            if new_barcode and new_barcode != old_barcode:
                if self.is_valid_barcode(new_barcode, barcode_type='item'):
                    if new_barcode not in self.all_boxes[box_barcode]:
                        self.undo_manager.add_action({
                            'type': 'edit_barcode',
                            'barcode': new_barcode,
                            'box_barcode': box_barcode,
                            'old_value': old_barcode,
                            'new_value': new_barcode,
                            'count': count
                        })
                        
                        self.all_boxes[box_barcode].pop(old_barcode)
                        self.all_boxes[box_barcode][new_barcode] = count
                        
                        if (box_barcode, old_barcode) in self.comments:
                            self.comments[(box_barcode, new_barcode)] = self.comments.pop((box_barcode, old_barcode))
                        
                        self.scan_history.append({
                            'timestamp': datetime.now().isoformat(),
                            'type': 'item',
                            'barcode': new_barcode,
                            'box_barcode': box_barcode,
                            'action': 'edit_barcode',
                            'action_type': 'edit',
                            'details': f'{old_barcode} → {new_barcode}'
                        })
                        
                        self.has_unsaved_changes = True
                        self.refresh_treeview()
                        self.update_undo_button_state()
                        self.scan_notification.show_notification(f"✅ Штрихкод изменён")
                    else:
                        self.show_error("Товар с таким штрихкодом уже есть в этом коробе!")
                else:
                    self.show_error("Неверный штрихкод товара!")

    def delete_box(self, item):
        box_barcode = item.text(1)
        dialog = ConfirmationDialog(
            "🗑️ Подтверждение удаления",
            f"Вы уверены, что хотите удалить короб '{box_barcode}'?",
            "warning",
            self
        )
        if dialog.exec_() == QDialog.Accepted:
            
            self.scan_history.append({
                'timestamp': datetime.now().isoformat(),
                'type': 'box',
                'barcode': box_barcode,
                'action': 'delete',
                'action_type': 'edit',
                'details': f'Удаление короба'
            })
            
            del self.all_boxes[box_barcode]
            keys_to_delete = []
            for key in self.comments:
                if key[0] == box_barcode:
                    keys_to_delete.append(key)
            for key in keys_to_delete:
                del self.comments[key]
            if self.current_box_barcode == box_barcode:
                self.current_box_barcode = ""
                self.update_status("")
            
            self.has_unsaved_changes = True
            self.refresh_treeview()

    def delete_item(self, item):
        parent_item = item.parent()
        box_barcode = parent_item.text(1) if parent_item else ""
        item_barcode = item.text(2)

        dialog = ConfirmationDialog(
            "🗑️ Подтверждение удаления",
            f"Вы уверены, что хотите удалить товар '{item_barcode}' из короба '{box_barcode}'?",
            "warning",
            self
        )
        if dialog.exec_() == QDialog.Accepted:
            
            self.scan_history.append({
                'timestamp': datetime.now().isoformat(),
                'type': 'item',
                'barcode': item_barcode,
                'box_barcode': box_barcode,
                'action': 'delete',
                'action_type': 'edit',
                'details': f'Удаление товара'
            })
            
            del self.all_boxes[box_barcode][item_barcode]
            if (box_barcode, item_barcode) in self.comments:
                del self.comments[(box_barcode, item_barcode)]
            if not self.all_boxes[box_barcode]:
                del self.all_boxes[box_barcode]
            if (box_barcode, "") in self.comments:
                del self.comments[(box_barcode, "")]
            if self.current_box_barcode == box_barcode:
                self.current_box_barcode = ""
                self.update_status("")
            
            self.has_unsaved_changes = True
            self.refresh_treeview()

    def edit_comment(self, item):
        values = [item.text(i) for i in range(self.items_tree.columnCount())]
    
        if len(values) == 6 and values[2] == "" and values[3] == "" and values[4] == "":
            # Для короба
            box_barcode = values[1]
            current_comment = self.comments.get((box_barcode, ""), "")
        
            # Создаем свой диалог для короба
            dialog = QDialog(self)
            dialog.setWindowTitle("✏️ Комментарий для короба")
            dialog.setModal(True)
            dialog.setMinimumWidth(350)
            dialog.setMaximumHeight(250)
            dialog.setWindowFlags(Qt.Dialog | Qt.WindowCloseButtonHint)
        
            layout = QVBoxLayout(dialog)
            layout.setSpacing(10)
            layout.setContentsMargins(15, 15, 15, 15)
        
            # Информация о коробе
            info_label = QLabel(f"Короб: {box_barcode}")
            info_label.setStyleSheet("""
                QLabel {
                    background-color: #f0f0f0;
                    padding: 8px;
                    border-radius: 4px;
                    font-weight: bold;
                    color: #2c3e50;
                }
            """)
            info_label.setWordWrap(True)
            layout.addWidget(info_label)
        
            # Поле для ввода комментария
            comment_edit = QTextEdit()
            comment_edit.setPlainText(current_comment)
            comment_edit.setMinimumHeight(100)
            comment_edit.setStyleSheet("""
                QTextEdit {
                    border: 2px solid #3498db;
                    border-radius: 4px;
                    padding: 5px;
                    font-size: 10pt;
                }
                QTextEdit:focus {
                    border-color: #e67e22;
                }
            """)
            layout.addWidget(comment_edit)
        
            # Кнопки OK/Отмена
            buttons_layout = QHBoxLayout()
            buttons_layout.setSpacing(10)
        
            ok_button = QPushButton("✅ Применить")
            ok_button.setFixedHeight(35)
            ok_button.setStyleSheet("""
                QPushButton {
                    background-color: #2ecc71;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    font-weight: bold;
                    font-size: 10pt;
                    padding: 0 15px;
                }
                QPushButton:hover {
                    background-color: #27ae60;
                }
            """)
            ok_button.clicked.connect(dialog.accept)
            buttons_layout.addWidget(ok_button)
        
            cancel_button = QPushButton("✕ Отмена")
            cancel_button.setFixedHeight(35)
            cancel_button.setStyleSheet("""
                QPushButton {
                    background-color: #e74c3c;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    font-weight: bold;
                    font-size: 10pt;
                    padding: 0 15px;
                }
                QPushButton:hover {
                    background-color: #c0392b;
                }
            """)
            cancel_button.clicked.connect(dialog.reject)
            buttons_layout.addWidget(cancel_button)
        
            layout.addLayout(buttons_layout)
        
            # Настройки
            comment_edit.setFocus()
        
            if dialog.exec_() == QDialog.Accepted:
                new_comment = comment_edit.toPlainText().strip()
                if new_comment != current_comment:
                    self.comments[(box_barcode, "")] = new_comment
                
                    self.scan_history.append({
                        'timestamp': datetime.now().isoformat(),
                        'type': 'box',
                        'barcode': box_barcode,
                        'action': 'edit_comment',
                        'action_type': 'edit',
                        'details': f'Комментарий: "{current_comment}" → "{new_comment}"'
                    })
                
                    self.has_unsaved_changes = True
                    self.refresh_treeview()
    
        else:
            # Для товара
            parent_item = item.parent()
            box_barcode = parent_item.text(1) if parent_item else ""
            item_barcode = values[2]
            current_comment = self.comments.get((box_barcode, item_barcode), "")
        
            # Создаем свой диалог для товара
            dialog = QDialog(self)
            dialog.setWindowTitle("✏️ Комментарий для товара")
            dialog.setModal(True)
            dialog.setMinimumWidth(350)
            dialog.setMaximumHeight(250)
            dialog.setWindowFlags(Qt.Dialog | Qt.WindowCloseButtonHint)
        
            layout = QVBoxLayout(dialog)
            layout.setSpacing(10)
            layout.setContentsMargins(15, 15, 15, 15)
        
            # Информация о товаре
            info_label = QLabel(f"Товар: {item_barcode}\nКороб: {box_barcode}")
            info_label.setStyleSheet("""
                QLabel {
                    background-color: #f0f0f0;
                    padding: 8px;
                    border-radius: 4px;
                    font-weight: bold;
                    color: #2c3e50;
                }
            """)
            info_label.setWordWrap(True)
            layout.addWidget(info_label)
        
            # Поле для ввода комментария
            comment_edit = QTextEdit()
            comment_edit.setPlainText(current_comment)
            comment_edit.setMinimumHeight(100)
            comment_edit.setStyleSheet("""
                QTextEdit {
                    border: 2px solid #3498db;
                    border-radius: 4px;
                    padding: 5px;
                    font-size: 10pt;
                }
                QTextEdit:focus {
                    border-color: #e67e22;
                }
            """)
            layout.addWidget(comment_edit)
        
            # Кнопки OK/Отмена
            buttons_layout = QHBoxLayout()
            buttons_layout.setSpacing(10)
            
            ok_button = QPushButton("✅ Применить")
            ok_button.setFixedHeight(35)
            ok_button.setStyleSheet("""
                QPushButton {
                    background-color: #2ecc71;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    font-weight: bold;
                    font-size: 10pt;
                    padding: 0 15px;
                }
               QPushButton:hover {
                    background-color: #27ae60;
                }
            """)
            ok_button.clicked.connect(dialog.accept)
            buttons_layout.addWidget(ok_button)
        
            cancel_button = QPushButton("✕ Отмена")
            cancel_button.setFixedHeight(35)
            cancel_button.setStyleSheet("""
                QPushButton {
                    background-color: #e74c3c;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    font-weight: bold;
                    font-size: 10pt;
                    padding: 0 15px;
                }
                QPushButton:hover {
                    background-color: #c0392b;
                }
            """)
            cancel_button.clicked.connect(dialog.reject)
            buttons_layout.addWidget(cancel_button)
        
            layout.addLayout(buttons_layout)
        
            # Настройки
            comment_edit.setFocus()
        
            if dialog.exec_() == QDialog.Accepted:
                new_comment = comment_edit.toPlainText().strip()
                if new_comment != current_comment:
                    self.comments[(box_barcode, item_barcode)] = new_comment
                
                    self.scan_history.append({
                        'timestamp': datetime.now().isoformat(),
                        'type': 'item',
                        'barcode': item_barcode,
                        'box_barcode': box_barcode,
                        'action': 'edit_comment',
                        'action_type': 'edit',
                        'details': f'Комментарий: "{current_comment}" → "{new_comment}"'
                    })
                
                    self.has_unsaved_changes = True
                    self.refresh_treeview()

    def on_double_click(self, item, column_index):
        if column_index in [3]:
            self.edit_item_count(item)
    
    def save_with_format_dialog(self):
        if not self.all_boxes:
            self.show_warning("Нет данных для сохранения!")
            return
            
        dialog = SaveFormatDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            format_index = dialog.get_selected_format()
            
            if format_index == 0:
                self.save_to_csv()
            elif format_index == 1:
                self.save_to_excel_multi_sheet()
            elif format_index == 2:
                self.save_to_excel_single_sheet()
    
    def save_to_excel_multi_sheet(self):
        if not self.all_boxes:
            self.show_warning("Нет данных для сохранения!")
            return

        file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить в Excel (много листов)", "", "Excel Files (*.xlsx);;All Files (*)")
        if not file_path:
            return
        if not file_path.lower().endswith(('.xlsx')):
            file_path += '.xlsx'

        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for box_barcode, items in self.all_boxes.items():
                sheet = wb.create_sheet(title=f"Короб {box_barcode[:15]}")
                
                if self.packer_name:
                    sheet['A1'] = "Сборщик:"
                    sheet['B1'] = self.packer_name
                    sheet['A1'].font = Font(bold=True)
                    sheet['B1'].font = Font(bold=True)
                
                sheet['A3'] = "Штрихкод короба"
                sheet['B3'] = box_barcode
                sheet['C3'] = "Комментарий"
                
                sheet['A4'] = "Штрихкод товара"
                sheet['B4'] = "Количество"
                sheet['C4'] = "Комментарий"
                if self.invoice_loaded:
                    sheet['D4'] = "План"
                    sheet['E4'] = "Статус"
                
                for cell in ['A3', 'B3', 'C3', 'A4', 'B4', 'C4']:
                    if cell in sheet:
                        sheet[cell].alignment = Alignment(horizontal='center')
                        sheet[cell].font = header_font
                        sheet[cell].fill = header_fill
                        sheet[cell].border = thin_border
                
                if self.invoice_loaded:
                    sheet['D4'].alignment = Alignment(horizontal='center')
                    sheet['D4'].font = header_font
                    sheet['D4'].fill = header_fill
                    sheet['D4'].border = thin_border
                    sheet['E4'].alignment = Alignment(horizontal='center')
                    sheet['E4'].font = header_font
                    sheet['E4'].fill = header_fill
                    sheet['E4'].border = thin_border
                
                row = 5
                sheet.cell(row=row, column=1, value="Комментарий к коробу:")
                sheet.cell(row=row, column=3, value=self.comments.get((box_barcode, ""), ""))
                sheet[f'A{row}'].font = Font(bold=True)
                row += 1
                
                for item_barcode, count in items.items():
                    sheet.cell(row=row, column=1, value=item_barcode)
                    sheet.cell(row=row, column=2, value=count).alignment = Alignment(horizontal='center')
                    sheet.cell(row=row, column=3, value=self.comments.get((box_barcode, item_barcode), ""))
                    
                    if self.invoice_loaded and item_barcode in self.invoice_data:
                        planned = self.invoice_data[item_barcode]
                        sheet.cell(row=row, column=4, value=planned).alignment = Alignment(horizontal='center')
                        
                        total_scanned = self.get_total_scanned_for_item(item_barcode)
                        if total_scanned == planned:
                            status = "✅ Совпадает"
                        elif total_scanned < planned:
                            status = f"⚠️ Недобор (план: {planned}, всего: {total_scanned}, не хватает: {planned - total_scanned})"
                        else:
                            status = f"❗ Перебор (план: {planned}, всего: {total_scanned}, лишних: {total_scanned - planned})"
                        sheet.cell(row=row, column=5, value=status)
                    elif self.invoice_loaded:
                        sheet.cell(row=row, column=4, value="0")
                        sheet.cell(row=row, column=5, value="❓ Лишний")
                    
                    for col in range(1, 6 if self.invoice_loaded else 4):
                        cell = sheet.cell(row=row, column=col)
                        if cell.value:
                            cell.border = thin_border
                    
                    row += 1

                for column in sheet.columns:
                    max_length = 0
                    col_letter = openpyxl.utils.get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    sheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
            wb.save(file_path)
            
            csv_path = self.save_csv_auto(file_path)
            if csv_path:
                self.has_unsaved_changes = False
                self.show_info(f"✅ Excel сохранен: {os.path.basename(file_path)}\n✅ Лог CSV сохранен: {os.path.basename(csv_path)}")
            else:
                self.show_warning(f"⚠️ Excel сохранен, но не удалось сохранить лог CSV!\n{os.path.basename(file_path)}")
        
        except Exception as e:
            self.show_error(f"Ошибка при сохранении: {e}")

    def save_to_excel_single_sheet(self):
        if not self.all_boxes:
            self.show_warning("Нет данных для сохранения!")
            return

        file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить в Excel (один лист)", "", "Excel Files (*.xlsx);;All Files (*)")
        if not file_path:
            return
        if not file_path.lower().endswith(('.xlsx')):
            file_path += '.xlsx'

        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Сборка"
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            if self.packer_name:
                sheet['A1'] = "Сборщик:"
                sheet['B1'] = self.packer_name
                sheet['A1'].font = Font(bold=True)
                sheet['B1'].font = Font(bold=True)
                sheet['A1'].fill = header_fill
                sheet['B1'].fill = header_fill
            
            sheet['A3'] = "Штрихкод короба"
            sheet['B3'] = "Комментарий короба"
            sheet['C3'] = "Штрихкод товара"
            sheet['D3'] = "Количество"
            sheet['E3'] = "Комментарий товара"
            
            if self.invoice_loaded:
                sheet['F3'] = "План"
                sheet['G3'] = "Статус"
            
            for col in range(1, 8 if self.invoice_loaded else 6):
                cell = sheet.cell(row=3, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            
            row = 4
            for box_barcode, items in self.all_boxes.items():
                box_comment = self.comments.get((box_barcode, ""), "")
                
                first_in_box = True
                for item_barcode, count in items.items():
                    item_comment = self.comments.get((box_barcode, item_barcode), "")
                    
                    if first_in_box:
                        sheet.cell(row=row, column=1, value=box_barcode)
                        sheet.cell(row=row, column=2, value=box_comment)
                        first_in_box = False
                    
                    sheet.cell(row=row, column=3, value=item_barcode)
                    sheet.cell(row=row, column=4, value=count).alignment = Alignment(horizontal='center')
                    sheet.cell(row=row, column=5, value=item_comment)
                    
                    if self.invoice_loaded and item_barcode in self.invoice_data:
                        planned = self.invoice_data[item_barcode]
                        sheet.cell(row=row, column=6, value=planned).alignment = Alignment(horizontal='center')
                        
                        total_scanned = self.get_total_scanned_for_item(item_barcode)
                        if total_scanned == planned:
                            status = "✅ Совпадает"
                        elif total_scanned < planned:
                            status = f"⚠️ Недобор (план: {planned}, всего: {total_scanned}, не хватает: {planned - total_scanned})"
                        else:
                            status = f"❗ Перебор (план: {planned}, всего: {total_scanned}, лишних: {total_scanned - planned})"
                        sheet.cell(row=row, column=7, value=status)
                    elif self.invoice_loaded:
                        sheet.cell(row=row, column=6, value="0")
                        sheet.cell(row=row, column=7, value="❓ Лишний")
                    
                    for col in range(1, 8 if self.invoice_loaded else 6):
                        cell = sheet.cell(row=row, column=col)
                        if cell.value:
                            cell.border = thin_border
                    
                    row += 1
                
                if items:
                    row += 1

            for column in sheet.columns:
                max_length = 0
                col_letter = openpyxl.utils.get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                sheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
            wb.save(file_path)
            
            csv_path = self.save_csv_auto(file_path)
            if csv_path:
                self.has_unsaved_changes = False
                self.show_info(f"✅ Excel сохранен: {os.path.basename(file_path)}\n✅ Лог CSV сохранен: {os.path.basename(csv_path)}")
            else:
                self.show_warning(f"⚠️ Excel сохранен, но не удалось сохранить лог CSV!\n{os.path.basename(file_path)}")
        
        except Exception as e:
            self.show_error(f"Ошибка при сохранении: {e}")
    
    def save_csv_auto(self, excel_path):
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_name = os.path.splitext(os.path.basename(excel_path))[0]
            csv_dir = os.path.dirname(excel_path)
            csv_path = os.path.join(csv_dir, f"{base_name}_{timestamp}.csv")
            
            with open(csv_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Сборщик", "Штрихкод короба", "Комментарий короба", "Штрихкод товара", "Количество", "Комментарий товара", "Время сканирования короба", "Время сканирования товара", "Тип действия", "Детали"])
                
                for box_barcode, items in self.all_boxes.items():
                    box_comment = self.comments.get((box_barcode, ""), "")
                    
                    box_timestamp = ""
                    for entry in self.scan_history:
                        if entry['type'] == 'box' and entry['barcode'] == box_barcode:
                            try:
                                dt = datetime.fromisoformat(entry['timestamp'])
                                box_timestamp = dt.strftime("%d.%m.%Y %H:%M:%S")
                            except:
                                box_timestamp = entry['timestamp']
                            break
                    
                    for item_barcode, count in items.items():
                        item_comment = self.comments.get((box_barcode, item_barcode), "")
                        
                        item_entries = []
                        for entry in self.scan_history:
                            if (entry['type'] == 'item' and 
                                entry['barcode'] == item_barcode and 
                                entry.get('box_barcode') == box_barcode):
                                try:
                                    dt = datetime.fromisoformat(entry['timestamp'])
                                    item_entries.append((
                                        dt.strftime("%d.%m.%Y %H:%M:%S"), 
                                        entry.get('action_type', 'scan'),
                                        entry.get('action', 'scan'),
                                        entry.get('details', '')
                                    ))
                                except:
                                    item_entries.append((
                                        entry['timestamp'],
                                        entry.get('action_type', 'scan'),
                                        entry.get('action', 'scan'),
                                        entry.get('details', '')
                                    ))
                        
                        # Записываем финальное количество одной строкой
                        writer.writerow([self.packer_name, box_barcode, box_comment, item_barcode, count, item_comment, box_timestamp, "", "final", f"Итоговое количество: {count}"])
                        
                        # Записываем историю изменений отдельно
                        for item_time, action_type, action, details in item_entries:
                            if action_type != 'final':
                                writer.writerow([self.packer_name, box_barcode, box_comment, item_barcode, 1, item_comment, box_timestamp, item_time, action_type, details])
            
            return csv_path
                            
        except Exception as e:
            print(f"Ошибка при сохранении CSV: {e}")
            return None
    
    def save_to_csv(self):
        if not self.all_boxes:
           self.show_warning("Нет данных для сохранения!")
           return
        file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить в CSV", "", "CSV Files (*.csv);;All Files (*)")
        if not file_path:
            return
        if not file_path.lower().endswith(('.csv')):
            file_path += '.csv'

        try:
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Сборщик", "Штрихкод короба", "Комментарий короба", "Штрихкод товара", "Количество", "Комментарий товара", "Время сканирования короба", "Время сканирования товара", "Тип действия", "Детали"])

                for box_barcode, items in self.all_boxes.items():
                    box_comment = self.comments.get((box_barcode, ""), "")
                    
                    box_timestamp = ""
                    for entry in self.scan_history:
                        if entry['type'] == 'box' and entry['barcode'] == box_barcode:
                            try:
                                dt = datetime.fromisoformat(entry['timestamp'])
                                box_timestamp = dt.strftime("%d.%m.%Y %H:%M:%S")
                            except:
                                box_timestamp = entry['timestamp']
                            break
                    
                    for item_barcode, count in items.items():
                        item_comment = self.comments.get((box_barcode, item_barcode), "")
                        
                        item_entries = []
                        for entry in self.scan_history:
                            if (entry['type'] == 'item' and 
                                entry['barcode'] == item_barcode and 
                                entry.get('box_barcode') == box_barcode):
                                try:
                                    dt = datetime.fromisoformat(entry['timestamp'])
                                    item_entries.append((
                                        dt.strftime("%d.%m.%Y %H:%M:%S"), 
                                        entry.get('action_type', 'scan'),
                                        entry.get('action', 'scan'),
                                        entry.get('details', '')
                                    ))
                                except:
                                    item_entries.append((
                                        entry['timestamp'],
                                        entry.get('action_type', 'scan'),
                                        entry.get('action', 'scan'),
                                        entry.get('details', '')
                                    ))
                        
                        # Записываем финальное количество одной строкой
                        writer.writerow([self.packer_name, box_barcode, box_comment, item_barcode, count, item_comment, box_timestamp, "", "final", f"Итоговое количество: {count}"])
                        
                        # Записываем историю изменений отдельно
                        for item_time, action_type, action, details in item_entries:
                            if action_type != 'final':
                                writer.writerow([self.packer_name, box_barcode, box_comment, item_barcode, 1, item_comment, box_timestamp, item_time, action_type, details])

            self.has_unsaved_changes = False
            self.show_info(f"✅ Данные сохранены в {file_path}")
        except Exception as e:
            self.show_error(f"Ошибка при сохранении: {e}")

    def load_invoice_dialog(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Загрузить накладную Excel", "", "Excel Files (*.xlsx *.xls);;All Files (*)")
        if file_path:
            self.show_loader(self._load_invoice_task, file_path)
    
    def _load_invoice_task(self, file_path, progress_callback=None, status_callback=None):
        import time
        start_time = time.time()
        
        if status_callback:
            status_callback("📂 Чтение файла Excel...")
    
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
    
        total_rows = sheet.max_row - 1
    
        if total_rows > 0 and progress_callback:
           progress_callback(0, total_rows)
           time.sleep(0.1)  # Пауза чтобы увидеть 0%
    
        invoice_data = {}
        total_items = 0
        total_quantity = 0
    
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
            if progress_callback:
                progress_callback(i, total_rows)
            
            if i % 5 == 0 and status_callback:
                percent = int((i / total_rows) * 100)
                status_callback(f"📊 Загружено {percent}% ({i}/{total_rows})")
                time.sleep(0.02)  # Маленькая задержка для плавности
        
            if row[0] and row[1]:
                barcode = str(row[0]).strip()
                try:
                    count = int(float(row[1]))
                    if count > 0:
                        invoice_data[barcode] = count
                        total_items += 1
                        total_quantity += count
                except:
                    continue
    
        min_animation_time = 1.5
        elapsed = time.time() - start_time
        
        if elapsed < min_animation_time and progress_callback:
            remaining = min_animation_time - elapsed
            steps = int(remaining * 10)
            
            for step in range(steps):
                progress_callback(total_rows, total_rows)
                if status_callback:
                    status_callback(f"✅ Финальная обработка... {int((step/steps)*100)}%")
                time.sleep(0.1)
    
        if progress_callback:
            progress_callback(total_rows, total_rows)
        if status_callback:
            status_callback("✅ Загрузка завершена")
    
        return (invoice_data, total_items, total_quantity, os.path.basename(file_path), file_path)
    
    def view_invoice(self):
        if not self.invoice_loaded or not self.invoice_data:
            return
        
        dialog = InvoiceViewDialog(self.invoice_data, self.invoice_file_name, self)
        dialog.exec_()
    
    def clear_invoice(self):
        if not self.invoice_loaded:
            return
            
        dialog = ConfirmationDialog(
            "⚠️ Сброс накладной",
            "Вы уверены, что хотите сбросить накладную?\n\nСтатистика (✅⚠️❗) будет сброшена, но собранные товары останутся в коробах.",
            "warning",
            self
        )
        dialog.yes_button.setText("✅ Да, сбросить")
        dialog.no_button.setText("✕ Нет")
        
        if dialog.exec_() == QDialog.Accepted:
            self.invoice_data = {}
            self.invoice_loaded = False
            self.invoice_file_name = ""
            self.invoice_file_path = ""
            self.invoice_label.setText("")
            self.clear_invoice_button.setEnabled(False)
            self.view_invoice_button.setEnabled(False)
            self.pause_button.hide()
            self.start_time = None
            self.first_scan_done = False
            self.is_paused = False
            self.match_label.setText("✅ 0")
            self.shortage_label.setText("⚠️ 0 (0)")
            self.excess_label.setText("❗ 0 (+0)")
            self.extra_label.setText("❓ 0 (+0)")
            self.progress_bar.setValue(0)
            self.time_label.setText("⏱️ Время: 00:00:00")
            self.speed_label.setText("⚡ Скорость: 0/мин")
            self.refresh_treeview()
            self.update_status("Накладная сброшена")
            self.status_bar.showMessage("💡 Перетащите CSV или Excel файл в окно для быстрого импорта")

    def load_from_csv_dialog(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Загрузить из CSV", "", "CSV Files (*.csv);;All Files (*)")
        if file_path:
            self._drag_import_file = file_path
            self.show_loader(self._load_csv_task, file_path)
    
    def _load_csv_task(self, file_path, progress_callback=None, status_callback=None):
        import time
        start_time = time.time()
        
        if status_callback:
            status_callback("📂 Чтение CSV файла...")
    
        with open(file_path, "r", encoding="utf-8") as f:
            all_lines = f.readlines()
    
        total_rows = len(all_lines) - 1
        if total_rows <= 0:
            raise Exception("Файл пуст")
    
        if progress_callback:
            progress_callback(0, total_rows)
            time.sleep(0.1)
    
        with open(file_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader, None)
        
            has_packer = len(header) >= 1 and header[0] == "Сборщик"
            has_timestamps = len(header) >= 8 and header[6] == "Время сканирования короба" and header[7] == "Время сканирования товара"
            has_action_types = len(header) >= 10 and header[8] == "Тип действия" and header[9] == "Детали"

            if not (len(header) >= 4 and header[1] == "Штрихкод короба" and header[3] == "Штрихкод товара" and header[4] == "Количество"):
                raise Exception("Некорректный формат файла CSV")

            all_boxes = {}
            comments = {}
            scan_history = []
            packer_name = ""
            box_timestamps = {}
            
            # Сначала собираем все действия в хронологическом порядке
            actions = []
        
            for row_idx, row in enumerate(reader, 1):
                if progress_callback:
                    progress_callback(row_idx, total_rows)
                
                if row_idx % 50 == 0 and status_callback:
                    percent = int((row_idx / total_rows) * 100)
                    status_callback(f"📊 Загружено {percent}% ({row_idx}/{total_rows})")
                    time.sleep(0.02)
            
                if len(row) < 5:
                    continue

                col_offset = 1 if has_packer else 0
            
                if has_packer and len(row) > 0 and row[0] and not packer_name:
                    packer_name = row[0]
            
                box_barcode = row[col_offset].strip() if len(row) > col_offset else ""
                box_comment = row[col_offset + 1].strip() if len(row) > col_offset + 1 else ""
                item_barcode = row[col_offset + 2].strip() if len(row) > col_offset + 2 else ""
                count_str = row[col_offset + 3].strip() if len(row) > col_offset + 3 else ""
                item_comment = row[col_offset + 4].strip() if len(row) > col_offset + 4 else ""
            
                box_timestamp = row[col_offset + 5].strip() if has_timestamps and len(row) > col_offset + 5 else ""
                item_timestamp = row[col_offset + 6].strip() if has_timestamps and len(row) > col_offset + 6 else ""
                action_type = row[col_offset + 7].strip() if has_action_types and len(row) > col_offset + 7 else "scan"
                details = row[col_offset + 8].strip() if has_action_types and len(row) > col_offset + 8 else ""

                if not box_barcode or not item_barcode:
                    continue

                if not self.is_valid_barcode(box_barcode, barcode_type='box'):
                    continue
                if not self.is_valid_barcode(item_barcode, barcode_type='item'):
                    continue
                
                try:
                    count = int(count_str)
                    if count <= 0:
                        continue
                except ValueError:
                    continue

                # Инициализация короба
                if box_barcode not in all_boxes:
                    all_boxes[box_barcode] = {}
                
                if box_timestamp and box_barcode not in box_timestamps:
                    # Преобразуем время для сортировки
                    try:
                        dt = datetime.strptime(box_timestamp, "%d.%m.%Y %H:%M:%S")
                        ts = dt.timestamp()
                    except:
                        ts = 0
                    box_timestamps[box_barcode] = (box_timestamp, action_type, details, ts)

                # Сохраняем комментарии
                comments[(box_barcode, "")] = box_comment
                comments[(box_barcode, item_barcode)] = item_comment
            
                # Сохраняем действие для последующего воспроизведения
                if item_timestamp:
                    try:
                        dt = datetime.strptime(item_timestamp, "%d.%m.%Y %H:%M:%S")
                        ts = dt.timestamp()
                        iso_timestamp = dt.isoformat()
                    except:
                        ts = 0
                        iso_timestamp = item_timestamp
                    
                    actions.append({
                        'timestamp': iso_timestamp,
                        'timestamp_float': ts,
                        'type': 'item',
                        'barcode': item_barcode,
                        'box_barcode': box_barcode,
                        'action_type': action_type,
                        'details': details,
                        'count': count,
                        'original_count': count
                    })

            # Сортируем действия по времени
            actions.sort(key=lambda x: x.get('timestamp_float', 0))

            # Воспроизводим действия для получения финального состояния
            final_counts = {}
            
            for action in actions:
                key = (action['box_barcode'], action['barcode'])
                
                if key not in final_counts:
                    final_counts[key] = 0
                
                if action['action_type'] == 'scan':
                    final_counts[key] += 1
                elif action['action_type'] == 'edit' and '→' in action['details']:
                    # Изменение количества
                    try:
                        parts = action['details'].split('→')
                        new_part = parts[1].split('(')[0].strip()
                        new_val = int(new_part)
                        final_counts[key] = new_val
                    except:
                        pass
                elif action['action_type'] == 'undo':
                    if 'Отмена сканирования' in action['details']:
                        final_counts[key] = max(0, final_counts[key] - 1)
                    elif 'Отмена изменения количества' in action['details']:
                        # Парсим "10 → 5"
                        try:
                            parts = action['details'].split('→')
                            old_val = int(parts[1].split('(')[0].strip())
                            final_counts[key] = old_val
                        except:
                            pass
                elif action['action_type'] == 'final':
                    # Прямое указание финального количества
                    final_counts[key] = action['count']

            # Применяем финальные количества
            for (box_barcode, item_barcode), final_count in final_counts.items():
                if box_barcode in all_boxes:
                    all_boxes[box_barcode][item_barcode] = final_count

            # Добавляем записи о коробах в историю
            for box_barcode, (timestamp, action_type, details, ts) in box_timestamps.items():
                try:
                    dt = datetime.strptime(timestamp, "%d.%m.%Y %H:%M:%S")
                    iso_timestamp = dt.isoformat()
                except:
                    iso_timestamp = timestamp
            
                action = 'scan'
                if action_type == 'edit':
                    if 'изменение' in details.lower():
                        action = 'edit_barcode'
                    elif 'удаление' in details.lower():
                        action = 'delete'
        
                scan_history.append({
                    'timestamp': iso_timestamp,
                    'type': 'box',
                    'barcode': box_barcode,
                    'action': action,
                    'action_type': action_type,
                    'details': details
                })

            # Добавляем все действия в историю
            for action in actions:
                action_copy = action.copy()
                if 'timestamp_float' in action_copy:
                    del action_copy['timestamp_float']
                if 'original_count' in action_copy:
                    del action_copy['original_count']
                scan_history.append(action_copy)

            # Сортируем историю по времени
            scan_history.sort(key=lambda x: x.get('timestamp', ''))

            start_time_val = None
            first_scan_done = False
            if scan_history:
                try:
                    first_scan = min(scan_history, key=lambda x: x.get('timestamp', ''))
                    first_time = datetime.fromisoformat(first_scan['timestamp'])
                    start_time_val = first_time.timestamp()
                    first_scan_done = True
                except:
                    pass

            # Анимация завершения
            min_animation_time = 1.5
            elapsed = time.time() - start_time
            
            if elapsed < min_animation_time and progress_callback:
                remaining = min_animation_time - elapsed
                steps = int(remaining * 10)
                
                for step in range(steps):
                    progress_callback(total_rows, total_rows)
                    if status_callback:
                        status_callback(f"✅ Финальная обработка... {int((step/steps)*100)}%")
                    time.sleep(0.1)

            if progress_callback:
                progress_callback(total_rows, total_rows)
            if status_callback:
                status_callback("✅ Загрузка завершена")
    
            return (all_boxes, comments, scan_history, packer_name, start_time_val, first_scan_done, os.path.basename(file_path))

    def load_from_csv(self, progress_callback=None, status_callback=None):
        if hasattr(self, '_drag_import_file') and self._drag_import_file:
            file_path = self._drag_import_file
            self._drag_import_file = None
            self.show_loader(self._load_csv_task, file_path)

    def new_box(self):
        self.current_box_barcode = ""
        self.update_status("Введите штрихкод нового короба")
        self.box_entry.setEnabled(True)
        self.box_entry.clear()
        self.box_entry.setFocus()
        self.item_scan_entry.clear()
        self.item_scan_entry.setEnabled(False)
        self.update_undo_button_state()

    def reset_application(self):
        dialog = ConfirmationDialog(
            "🔄 Подтверждение",
            "Вы уверены, что хотите начать заново? Все несохранённые данные будут потеряны.",
            "warning",
            self
        )
        dialog.yes_button.setText("✅ Да")
        dialog.no_button.setText("✕ Нет")
    
        if dialog.exec_() == QDialog.Accepted:
            self.all_boxes = {}
            self.current_box_barcode = ""
            self.search_query = ""
            self.comments = {}
            self.scan_history = []
            self.undo_manager = UndoManager(max_size=10)
            self.packer_name = ""
            self.packer_combo.setCurrentText("")
            self.invoice_data = {}
            self.invoice_loaded = False
            self.invoice_file_name = ""
            self.invoice_file_path = ""
            self.invoice_label.setText("")
            self.clear_invoice_button.setEnabled(False)
            self.view_invoice_button.setEnabled(False)
            self.pause_button.hide()
            self.start_time = None
            self.first_scan_done = False
            self.is_paused = False
            self.total_scans = 0
            self.has_unsaved_changes = False
            self.box_entry.setEnabled(True)
            self.box_entry.clear()
            self.item_scan_entry.setEnabled(False)
            self.item_scan_entry.clear()
            self.search_entry.clear()
            self.match_label.setText("✅ 0")
            self.shortage_label.setText("⚠️ 0 (0)")
            self.excess_label.setText("❗ 0 (+0)")
            self.extra_label.setText("❓ 0 (+0)")
            self.progress_bar.setValue(0)
            self.time_label.setText("⏱️ Время: 00:00:00")
            self.speed_label.setText("⚡ Скорость: 0/мин")
            self.refresh_treeview()
            self.update_status("")
            self.box_entry.setFocus()
            self.save_button.setEnabled(False)
            self.update_undo_button_state()
            self.status_bar.showMessage("💡 Перетащите CSV или Excel файл в окно для быстрого импорта")
            self.save_state()

    def show_error(self, message):
        QMessageBox.critical(self, "❌ Ошибка", message)

    def show_warning(self, message):
        QMessageBox.warning(self, "⚠️ Предупреждение", message)

    def show_info(self, message):
        QMessageBox.information(self, "ℹ️ Информация", message)

    def update_status(self, message):
        self.status_bar.showMessage(message)

    def update_summary(self):
        num_boxes = len(self.all_boxes)
        total_items = 0
        for box, items in self.all_boxes.items():
            total_items += sum(items.values())
        
        if self.invoice_loaded:
            total_planned = sum(self.invoice_data.values())
            summary_text = f"📊 Коробов: {num_boxes} | Собрано: {total_items} | План: {total_planned}"
        else:
            summary_text = f"📊 Коробов: {num_boxes} | Товаров: {total_items}"
        
        self.summary_label.setText(summary_text)

    def load_state(self):
        try:
            if os.path.exists(self.state_file):
                with open(self.state_file, "r") as f:
                    data = json.load(f)
                    if 'all_boxes' in data:
                        self.all_boxes = {str(k): v for k, v in data['all_boxes'].items()}
                    if 'current_box_barcode' in data:
                        self.current_box_barcode = data['current_box_barcode']
                    if 'search_query' in data:
                        self.search_query = data['search_query']
                    if 'packer_name' in data:
                        self.packer_name = data['packer_name']
                        self.packer_combo.setCurrentText(self.packer_name)
                    if 'start_time' in data and data['start_time']:
                        self.start_time = data['start_time']
                        self.first_scan_done = True
                    if 'is_paused' in data:
                        self.is_paused = data['is_paused']
                    if 'total_scans' in data:
                        self.total_scans = data['total_scans']
                        
                    serializable_comments = data.get('comments', {})
                    self.comments = {}
                    for key_str, comment in serializable_comments.items():
                        try:
                            box_barcode, item_barcode_str = key_str.split(",", 1) if "," in key_str else (key_str, "")
                            item_barcode = item_barcode_str if item_barcode_str else ""
                            self.comments[(box_barcode, item_barcode)] = comment
                        except ValueError:
                            pass
                    if 'strict_validation_enabled' in data:
                        self.strict_validation_enabled = data['strict_validation_enabled']
                        if hasattr(self, 'strict_validation_checkbox'):
                            self.strict_validation_checkbox.setChecked(self.strict_validation_enabled)
                    
                    if 'scan_history' in data:
                        self.scan_history = data['scan_history']
                        
                self.refresh_treeview()
                if self.current_box_barcode:
                    self.box_entry.setEnabled(False)
                    self.item_scan_entry.setEnabled(True)
                    self.save_button.setEnabled(True)
                    
                if self.invoice_loaded:
                    self.pause_button.show()
                    if self.is_paused:
                        self.pause_button.setText("▶️")
                    else:
                        self.pause_button.setText("⏸️")
        except Exception as e:
            pass

    def save_state(self):
        serializable_comments = {}
        for key, comment in self.comments.items():
            if not isinstance(key, tuple) or len(key) != 2:
                continue
            box_barcode, item_barcode = key
            key_str = f"{box_barcode},{item_barcode}"
            serializable_comments[key_str] = comment
            
        data = {
            "all_boxes": self.all_boxes,
            "current_box_barcode": self.current_box_barcode,
            "search_query": self.search_query,
            "comments": serializable_comments,
            "strict_validation_enabled": self.strict_validation_enabled,
            "scan_history": self.scan_history,
            "packer_name": self.packer_name,
            "start_time": self.start_time,
            "is_paused": self.is_paused,
            "total_scans": self.total_scans,
        }
        try:
            with open(self.state_file, "w") as f:
                json.dump(data, f)
        except Exception as e:
            pass
            
    def load_column_settings(self):
        for i in range(6):
            default_width = 60 if i == 0 else 180 if i < 3 else 80 if i in (3, 4) else 200
            width = self.settings.value(f"column_width_{i}", default_width)
            self.items_tree.setColumnWidth(i, int(width))
            
    def save_column_settings(self):
        for i in range(6):
            self.settings.setValue(f"column_width_{i}", self.items_tree.columnWidth(i))

    def on_closing(self):
        self.save_column_settings()
        self.save_state()
        self.close()

    def show_paste_menu(self, event, entry_widget):
        context_menu = QMenu(self)
        paste_action = QAction("📋 Вставить", self)
        paste_action.triggered.connect(lambda: self.paste_from_clipboard(entry_widget))
        context_menu.addAction(paste_action)
        context_menu.popup(entry_widget.mapToGlobal(event))

    def paste_from_clipboard(self, entry_widget):
        text = self.clipboard.text()
        entry_widget.insert(text)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    barcode_app = QBarcodeApp()
    barcode_app.show()
    sys.exit(app.exec_())
